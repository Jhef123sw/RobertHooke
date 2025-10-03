import matplotlib
import zipfile
matplotlib.use('Agg')  # Forzar backend sin GUI
import matplotlib.pyplot as plt
import base64
import os
from datetime import time
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
import io
from django.urls import reverse
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest, FileResponse, Http404, JsonResponse
from django.template.loader import render_to_string
from django.contrib import messages
from datetime import datetime
import pandas as pd
import shutil
from collections import defaultdict
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from .models import Estudiante, Reporte, Asistencia, VariableControl
from .forms import EstudianteForm, CargarExcelForm, LoginForm, EstudianteForm2, CargarExcelFormReporte, ActualizarDatosForm, AsistenciaForm2, VariableControlForm, ActualizarDatosFormProf, ImagenPreguntaForm
from .backends import EstudianteBackend
from .decorators import estudiante_tipo_requerido, datos_actualizados_requerido
from django.core.paginator import Paginator
from django.db.models import Q
import matplotlib.pyplot as plt
import math
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, get_object_or_404
from .models import Reporte, Estudiante, curso
from datetime import datetime
from .tasks import generar_reportes_completos_task, generar_pdfs_asistencias_por_estudiante_task
from celery.result import AsyncResult
from weasyprint import HTML

#Cursos
@login_required
@estudiante_tipo_requerido(['administrador'])
def cursos(request):
    base_template = "layouts/base_profesor.html"
    data = []

    cursos = curso.objects.all()

    return render(request, 'cursos_por_profesor.html', {'cursos': cursos, 'base_template': base_template})

@login_required
@estudiante_tipo_requerido(['administrador'])
def editar_costo_hora(request):
    cursos = curso.objects.all()

    if request.method == "POST":
        for c in cursos:
            nuevo_costo = request.POST.get(f"costo_{c.id}")
            if nuevo_costo is not None:
                try:
                    c.costo_hora = float(nuevo_costo)
                    c.save()
                except ValueError:
                    messages.error(request, f"Costo inválido para {c.get_nombreCurso_display()}")
        messages.success(request, "Costos por hora actualizados correctamente ✅")
        return redirect("editar_costo_hora")

    return render(request, "editar_costo_hora.html", {'base_template': 'layouts/base.html', "cursos": cursos})






#Resultados
def dashboard_reportes2(request):
    return render(request, "dashboard_reportes2.html", {'base_template': 'layouts/base.html'})


def dashboard_reportes1(request):
    return render(request, "dashboard_reportes1.html", {'base_template': 'layouts/base.html'})

def dashboard_reportes(request):
    return render(request, "dashboard_reportes.html", {'base_template': 'layouts/base.html'})

def obtener_fechas_nivel(request, nivel):
    fechas = (
        Reporte.objects.filter(nivel=nivel)
        .order_by("-fecha_de_examen")
        .values_list("fecha_de_examen", flat=True)
        .distinct()
    )
    fechas_str = [f.strftime("%Y-%m-%d") for f in fechas]
    return JsonResponse({"fechas": fechas_str})

def obtener_resultados_fecha(request, nivel, fecha):
    reportes = Reporte.objects.filter(nivel=nivel, fecha_de_examen=fecha)

    resultados = []
    for r in reportes:
        resultados.append({
            "estudiante": str(r.KK_usuario.nombre),
            "datos": r.obtener_datos(),  # diccionario de materia: (buenas, malas)
            "puntaje": r.obtener_total_puntaje(),
            "puesto": r.puesto
        })

    return JsonResponse({"resultados": resultados})

def obtener_resultados_fecha1(request, nivel, fecha):
    reportes = Reporte.objects.filter(nivel=nivel, fecha_de_examen=fecha).order_by("puesto")
    if nivel == '90':
        total_preguntas = 90  # Ajusta según corresponda
    else:
        total_preguntas = 45  # Ajusta según corresponda
    resultados = []
    for r in reportes:
        datos = r.obtener_datos()
        buenas = sum(v[0] for v in datos.values())
        malas = sum(v[1] for v in datos.values())

        blancas = total_preguntas - (buenas + malas)

        porcentaje_buenas = round((buenas / total_preguntas) * 100, 2)
        porcentaje_malas = round((malas / total_preguntas) * 100, 2)
        porcentaje_blancas = round((blancas / total_preguntas) * 100, 2)
        puntaje = round(r.obtener_total_puntaje(), 2)

        resultados.append({
            "puesto": r.puesto,
            "usuario": str(r.KK_usuario.usuario),
            "estudiante": str(r.KK_usuario.nombre),
            "buenas": buenas,
            "malas": malas,
            "blancas": blancas,
            "porcentaje_buenas": porcentaje_buenas,
            "porcentaje_malas": porcentaje_malas,
            "porcentaje_blancas": porcentaje_blancas,
            "puntaje": puntaje
        })

    return JsonResponse({"resultados": resultados})


def obtener_resultados_tabla_todo(request, nivel, fecha):
    reportes = Reporte.objects.filter(nivel=nivel, fecha_de_examen=fecha).order_by("puesto")
    if nivel == '90':
        total_preguntas = 90  # Ajusta según corresponda
    else:
        total_preguntas = 45  # Ajusta según corresponda
    resultados = []
    for r in reportes:
        datos = r.obtener_datos()
        buenas = sum(v[0] for v in datos.values())
        malas = sum(v[1] for v in datos.values())

        blancas = total_preguntas - (buenas + malas)

        porcentaje_buenas = round((buenas / total_preguntas) * 100, 2)
        porcentaje_malas = round((malas / total_preguntas) * 100, 2)
        porcentaje_blancas = round((blancas / total_preguntas) * 100, 2)
        puntaje = round(r.obtener_total_puntaje(), 2)

        resultados.append({
            "puesto": r.puesto,
            "usuario": str(r.KK_usuario.usuario),
            "estudiante": str(r.KK_usuario.nombre),
            "buenas": buenas,
            "malas": malas,
            "blancas": blancas,
            "porcentaje_buenas": porcentaje_buenas,
            "porcentaje_malas": porcentaje_malas,
            "porcentaje_blancas": porcentaje_blancas,
            "puntaje": puntaje
        })

    return JsonResponse({"resultados": resultados})


def exportar_pdf(request, nivel, fecha):
    reportes = Reporte.objects.filter(nivel=nivel, fecha_de_examen=fecha).order_by("puesto")

    resultados = []
    total_preguntas = 90 if str(nivel) == "90" else 45
    for r in reportes:
        datos = r.obtener_datos()
        buenas = sum(v[0] for v in datos.values())
        malas = sum(v[1] for v in datos.values())
        blancas = total_preguntas - (buenas + malas)
        puntaje = round(r.obtener_total_puntaje(), 2)

        resultados.append({
            "puesto": r.puesto,
            "usuario": str(r.KK_usuario.usuario),
            "estudiante": str(r.KK_usuario.nombre),
            "buenas": buenas,
            "malas": malas,
            "blancas": blancas,
            "puntaje": puntaje
        })

    html_string = render_to_string("reportes_pdf.html", {
        "resultados": resultados,
        "nivel": "Pre" if str(nivel) == "90" else "Semillero",
        "fecha": fecha,
    })

    html = HTML(string=html_string)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="reporte_{fecha}.pdf"'
    html.write_pdf(response)
    return response



#Preguntas
def simulacros(request):
    return render(request, 'simulacros.html', {'base_template': 'layouts/base.html'})


def evaluacion_simulacros(request):
    return render(request, 'evaluacion_simulacros.html', {'base_template': 'layouts/base.html'})


def seleccionar_nivel(request):
    return render(request, 'seleccionar_nivel.html', {'base_template': 'layouts/base.html'})


def seleccionar_fecha(request, nivel):
    nivel_id = 90 if nivel == 'pre' else 30
    fechas = Reporte.objects.filter(nivel=nivel_id).order_by('fecha_de_examen').values_list('fecha_de_examen', flat=True).distinct()
    return render(request, 'seleccionar_fecha.html', {'nivel': nivel, 'fechas': fechas, 'base_template': 'layouts/base.html'})


def seleccionar_curso(request, nivel, fecha):
    cursos_pre = ['Raz_Verbal', 'Raz_Matemático', 'Aritmética', 'Álgebra', 'Geometría', 'Trigonometría', 'Física', 'Química', 'Biología', 'Lenguaje', 'Literatura', 'Historia', 'Geografía', 'Filosofía', 'Psicología', 'Economía']
    cursos_sem = ['Raz_Matemático','Raz_Verbal', 'Aritmética', 'Álgebra','Literatura']
    cursos = cursos_pre if nivel == 'pre' else cursos_sem

    carpeta = 'pre' if nivel == 'pre' else 'sem'
    ruta = os.path.join(settings.MEDIA_ROOT, f'preguntas/{carpeta}')
    
    cursos_con_imagen = []
    for curso in cursos:
        nombre_archivo = f"{nivel}_{fecha}_{curso}.png"
        ruta_completa = os.path.join(ruta, nombre_archivo)
        if os.path.exists(ruta_completa):
            cursos_con_imagen.append(curso)

    return render(request, 'seleccionar_curso.html', {
        'nivel': nivel,
        'fecha': fecha,
        'cursos': cursos,
        'cursos_con_imagen': cursos_con_imagen,
        'base_template': 'layouts/base.html'
    })


def subir_imagen_pregunta(request, nivel, fecha, curso):
    carpeta = 'pre' if nivel == 'pre' else 'sem'
    nombre_archivo = f"{nivel}_{fecha}_{curso}.png"
    ruta_relativa = f"preguntas/{carpeta}/{nombre_archivo}"
    ruta_absoluta = os.path.join(settings.MEDIA_ROOT, ruta_relativa)

    imagen_url = None
    if os.path.exists(ruta_absoluta):
        imagen_url = os.path.join(settings.MEDIA_URL, ruta_relativa)

    if request.method == 'POST':
        form = ImagenPreguntaForm(request.POST, request.FILES)
        if form.is_valid():
            imagen = form.cleaned_data['imagen']
            os.makedirs(os.path.dirname(ruta_absoluta), exist_ok=True)
            with open(ruta_absoluta, 'wb+') as destino:
                for chunk in imagen.chunks():
                    destino.write(chunk)
            return redirect('seleccionar_curso', nivel=nivel, fecha=fecha)
    else:
        form = ImagenPreguntaForm()

    return render(request, 'subir_imagen.html', {
        'form': form,
        'nivel': nivel,
        'fecha': fecha,
        'curso': curso,
        'imagen_url': imagen_url,
        'base_template': 'layouts/base.html'
    })


################Cursos-Profesores
@login_required
@estudiante_tipo_requerido(['administrador'])
def vista_grafico_por_tipo(request):
    return render(request, 'grafico_por_tipo.html', {'base_template': 'layouts/base.html'})

COLORES = [
    "red", "blue", "green", "orange", "purple", "brown", "pink",
    "teal", "gray", "lime", "navy", "maroon", "olive"
]

@csrf_exempt
@login_required
def obtener_grafico_por_tipo(request):
    nivel = request.GET.get('nivel')
    curso = request.GET.get('curso')
    tipo = request.GET.get('tipo')

    if not nivel or not curso or not tipo:
        return JsonResponse({'error': 'Parámetros insuficientes'}, status=400)

    nivel_id = 90 if nivel == 'Pre' else 30
    reportes = Reporte.objects.filter(nivel=nivel_id).order_by('fecha_de_examen')
    if not reportes.exists():
        return JsonResponse({'error': 'No hay reportes disponibles'}, status=404)

    config = VariableControl.objects.get(ID_Variable=1)
    total_preg = getattr(config, f"{curso}_Pre" if nivel_id == 90 else f"{curso}_Sem", 0)

    intervalos = INTERVALOS[nivel].get(curso, [])
    fechas_unicas = sorted(set(r.fecha_de_examen.strftime('%d/%m/%Y') for r in reportes))

    # Inicializar conteo: {intervalo: [0, 0, ...] por cada fecha}
    conteo_por_intervalo = {f"{ini}–{fin}": [0]*len(fechas_unicas) for (ini, fin) in intervalos}

    fecha_a_indice = {fecha: idx for idx, fecha in enumerate(fechas_unicas)}

    for reporte in reportes:
        fecha = reporte.fecha_de_examen.strftime('%d/%m/%Y')
        idx_fecha = fecha_a_indice[fecha]

        buenas = getattr(reporte, f"{curso}_1", 0)
        malas = getattr(reporte, f"{curso}_2", 0)
        blancas = max(0, total_preg - buenas - malas)

        valor = {
            'correctas': buenas,
            'incorrectas': malas,
            'blancas': blancas
        }.get(tipo, 0)

        for (ini, fin) in intervalos:
            if ini <= valor <= fin:
                clave = f"{ini}–{fin}"
                conteo_por_intervalo[clave][idx_fecha] += 1
                break

    # Preparar datos para Chart.js
    series = []
    for i, (clave, valores) in enumerate(conteo_por_intervalo.items()):
        series.append({
            'intervalo': clave,
            'valores': valores,
            'color': COLORES[i % len(COLORES)]
        })

    return JsonResponse({
        'fechas': fechas_unicas,
        'series': series
    })

def asistencia_desempeño_curso(request, id_curso):
    return render(request, 'asistencia_desempeño_curso.html', {'base_template': "layouts/base_profesor.html"})    


@login_required
@estudiante_tipo_requerido(['profesor'])
def cursos_por_profesor(request):
    base_template = "layouts/base_profesor.html"
    profe = request.user
    profesores = Estudiante.objects.filter(tipo_estudiante='profesor')
    data = []

    cursos = curso.objects.filter(estudiante = profe)

    for profesor in profesores:
        cursos_asignados = curso.objects.filter(estudiante=profesor)
        data.append({
            'profesor': profesor,
            'cursos': cursos_asignados,
        })

    return render(request, 'cursos_por_profesor.html', {'cursos': cursos, 'base_template': base_template})

INTERVALOS = {
    'Pre': {
        'Rv': [(0,2), (3,5), (6,8), (9,11), (12,15)],
        'Rm': [(0,2), (3,5), (6,8), (9,11), (12,15)],
        'Ar': [(0,), (1,), (2,), (3,), (4,), (5,)],
        'Al': [(0,), (1,), (2,), (3,), (4,), (5,)],
        'Ge': [(0,), (1,), (2,), (3,), (4,)],
        'Tr': [(0,), (1,), (2,), (3,), (4,)],
        'Fi': [(0,), (1,), (2,), (3,), (4,), (5,)],
        'Qu': [(0,), (1,), (2,), (3,), (4,), (5,)],
        'Bi': [(0,1), (2,3), (4,5), (6,7), (8,9)],
        'Le': [(0,), (1,), (2,), (3,), (4,), (5,6)],
        'Lit': [(0,), (1,), (2,), (3,), (4,), (5,6)],
        'Hi': [(0,), (1,), (2,), (3,)],
        'Gf': [(0,), (1,), (2,)],
        'Fil': [(0,), (1,), (2,)],
        'Psi': [(0,), (1,), (2,)],
        'Ec': [(0,), (1,), (2,)],
    },
    'Semillero': {
        'Rv': [(0,), (1,2), (3,4), (5,6), (7,8), (9,10)],
        'Rm': [(0,), (1,2), (3,4), (5,6), (7,8), (9,10)],
        'Ar': [(0,), (1,), (2,), (3,), (4,), (5,)],
        'Al': [(0,), (1,), (2,), (3,), (4,), (5,)],
        'Lit': [(0,), (1,), (2,), (3,), (4,), (5,)],
        # Los demás cursos con 0 preguntas no generarán datos
    }
}

@csrf_exempt
@login_required
def fechas_unicas_por_nivel_y_curso(request):
    nivel = request.GET.get('nivel')  # 'Pre' o 'Semillero'
    curso = request.GET.get('curso')  # Ej: 'Rv'

    if not nivel or not curso:
        return JsonResponse({'error': 'Parámetros insuficientes'}, status=400)

    nivel_id = 90 if nivel == 'Pre' else 30
    fechas = Reporte.objects.filter(nivel=nivel_id).order_by('fecha_de_examen').values_list('fecha_de_examen', flat=True).distinct()
    fechas_formateadas = [f.strftime('%d/%m/%Y') for f in fechas]

    return JsonResponse({'fechas': fechas_formateadas})



@login_required
def vista_grafico_intervalos(request):
    base_template = "layouts/base.html"
    return render(request, 'grafico_por_curso.html', {'base_template': base_template})

def api_grafico_curso(request):
    nivel = int(request.GET.get('nivel'))
    curso = request.GET.get('curso')

    estudiante = request.user
    reportes = Reporte.objects.filter(KK_usuario=estudiante, nivel=nivel).order_by('fecha_de_examen')
    fechas = [r.fecha_de_examen.strftime('%Y-%m-%d') for r in reportes]

    # Obtener correctas e incorrectas
    correctas = []
    incorrectas = []
    en_blanco = []

    config = VariableControl.objects.get(ID_Variable=1)
    clave = curso + ('_Pre' if nivel == 90 else '_Sem')
    total_preguntas = getattr(config, clave, 0)

    for r in reportes:
        datos = r.obtener_datos()
        nombre_curso = dict(Reporte._meta.get_field('Rv_1').choices).get(curso, curso)
        if nombre_curso not in datos:
            continue
        c, i = datos[nombre_curso]
        correctas.append(c)
        incorrectas.append(i)
        en_blanco.append(max(total_preguntas - (c + i), 0))

    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(fechas, correctas, label='Correctas', marker='o')
    ax.plot(fechas, incorrectas, label='Incorrectas', marker='o')
    ax.plot(fechas, en_blanco, label='En blanco', marker='o')

    ax.set_xlabel('Fecha')
    ax.set_ylabel('Cantidad')
    ax.set_title(f'Evolución por curso: {curso}')
    ax.legend()
    plt.xticks(rotation=90)
    fig.tight_layout()

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    buffer.close()
    plt.close(fig)

    return JsonResponse({'grafico_base64': imagen_base64})


@csrf_exempt
@login_required
def obtener_distribucion_por_fecha(request):
    nivel = request.GET.get('nivel')  # 'Pre' o 'Semillero'
    curso = request.GET.get('curso')  # Ej: 'Rv'
    fecha = request.GET.get('fecha')  # formato esperado: 'dd/mm/yyyy'

    if not nivel or not curso or not fecha:
        return JsonResponse({'error': 'Parámetros insuficientes'}, status=400)

    try:
        from datetime import datetime
        fecha_examen = datetime.strptime(fecha, '%d/%m/%Y').date()
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)

    nivel_id = 90 if nivel == 'Pre' else 30
    reportes = Reporte.objects.filter(nivel=nivel_id, fecha_de_examen=fecha_examen)

    if not reportes.exists():
        return JsonResponse({'error': 'No hay reportes disponibles para esa fecha'}, status=404)

    intervalos_raw = INTERVALOS[nivel].get(curso, [])
    intervalos_etiquetas = [f"{rango[0]}-{rango[-1]}" for rango in intervalos_raw]

    def obtener_indice_intervalo(valor):
        for i, intervalo in enumerate(intervalos_raw):
            start = intervalo[0]
            end = intervalo[-1] if len(intervalo) > 1 else intervalo[0]
            if start <= valor <= end:
                return i
        return None

    config = VariableControl.objects.get(ID_Variable=1)
    total_preguntas = getattr(config, f"{curso}_Pre" if nivel_id == 90 else f"{curso}_Sem", 0)

    # Contadores por intervalo
    correctas_contador = [0] * len(intervalos_raw)
    incorrectas_contador = [0] * len(intervalos_raw)
    blancas_contador = [0] * len(intervalos_raw)

    for reporte in reportes:
        correctas = getattr(reporte, f"{curso}_1", 0)
        incorrectas = getattr(reporte, f"{curso}_2", 0)
        blancas = max(0, total_preguntas - correctas - incorrectas)

        i_corr = obtener_indice_intervalo(correctas)
        i_inc = obtener_indice_intervalo(incorrectas)
        i_bla = obtener_indice_intervalo(blancas)

        if i_corr is not None:
            correctas_contador[i_corr] += 1
        if i_inc is not None:
            incorrectas_contador[i_inc] += 1
        if i_bla is not None:
            blancas_contador[i_bla] += 1

    return JsonResponse({
        'intervalos': intervalos_etiquetas,
        'correctas': correctas_contador,
        'incorrectas': incorrectas_contador,
        'blancas': blancas_contador,
    })


@csrf_exempt
@login_required
def obtener_grafico_por_intervalos(request):
    nivel = request.GET.get('nivel')  # 'Pre' o 'Semillero'
    curso = request.GET.get('curso')  # Ej: 'Rv'

    if not nivel or not curso:
        return JsonResponse({'error': 'Parámetros insuficientes'}, status=400)

    nivel_id = 90 if nivel == 'Pre' else 30
    reportes = Reporte.objects.filter(nivel=nivel_id).order_by('fecha_de_examen')

    if not reportes.exists():
        return JsonResponse({'error': 'No hay reportes disponibles'}, status=404)

    # Obtener intervalos definidos para el curso y nivel
    intervalos_raw = INTERVALOS[nivel].get(curso, [])
    intervalos_etiquetas = [f"{rango[0]}-{rango[-1]}" for rango in intervalos_raw]

    def obtener_indice_intervalo(valor):
        for i, intervalo in enumerate(intervalos_raw):
            start = intervalo[0]
            end = intervalo[-1] if len(intervalo) > 1 else intervalo[0]
            if start <= valor <= end:
                return i
        return None  # En caso de que el valor no encaje en ningún intervalo

    config = VariableControl.objects.get(ID_Variable=1)
    total_preguntas = getattr(config, f"{curso}_Pre" if nivel_id == 90 else f"{curso}_Sem", 0)

    fechas = []
    correctas_indices = []
    incorrectas_indices = []
    blancas_indices = []

    for reporte in reportes:
        fechas.append(reporte.fecha_de_examen.strftime('%d/%m/%Y'))

        correctas = getattr(reporte, f"{curso}_1", 0)
        incorrectas = getattr(reporte, f"{curso}_2", 0)
        blancas = max(0, total_preguntas - correctas - incorrectas)

        correctas_indices.append(obtener_indice_intervalo(correctas))
        incorrectas_indices.append(obtener_indice_intervalo(incorrectas))
        blancas_indices.append(obtener_indice_intervalo(blancas))

    return JsonResponse({
        'labels': fechas,
        'intervalos': intervalos_etiquetas,
        'correctas': correctas_indices,
        'incorrectas': incorrectas_indices,
        'blancas': blancas_indices,
    })
#################PROFESORES
@login_required
@estudiante_tipo_requerido(['administrador'])
def crear_profesor(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        usuario = request.POST.get('usuario')
        contraseña = request.POST.get('contraseña')

        if Estudiante.objects.filter(usuario=usuario).exists():
            return render(request, 'crear_tutor.html', {
                'mensaje': 'El nombre de usuario ya existe.'
            })

        Estudiante.objects.create(
            nombre=nombre,
            usuario=usuario,
            contraseña=contraseña,
            tipo_estudiante='profesor',
            is_active=True,
            is_staff=False,
        )
        return redirect('listar_profesores')

    return render(request, 'crear_tutor.html', {
        'base_template': base_template,
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def listar_profesores(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    profesores = Estudiante.objects.filter(tipo_estudiante='profesor')
    return render(request, 'listar_profesores.html', {
        'tutores': profesores,
        'base_template': base_template,
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def asignar_cursos_profesor(request):
    base_template = "layouts/base.html"
    profesores = Estudiante.objects.filter(tipo_estudiante='profesor')
    cursos = curso.objects.all()
    return render(request, 'asignar_cursos.html', {
        'base_template': base_template,
        'profesores': profesores,
        'cursos': cursos,
    })

@login_required
@estudiante_tipo_requerido(['administrador'])
def obtener_cursos_profesor(request, profesor_id):
    prof = get_object_or_404(Estudiante, pk=profesor_id, tipo_estudiante='profesor')
    cursos = prof.cursos.all().values('id', 'nombreCurso')
    return JsonResponse(list(cursos), safe=False)

@login_required
@estudiante_tipo_requerido(['administrador'])
def obtener_todos_los_cursos(request):
    cursos = curso.objects.all().select_related('estudiante')
    data = [{
        'id': c.id,
        'nombreCurso': c.get_nombreCurso_display(),
        'profesor': c.estudiante.usuario if c.estudiante else None,
    } for c in cursos]
    return JsonResponse(data, safe=False)

@login_required
@estudiante_tipo_requerido(['administrador'])
@require_POST
def asignar_curso(request):
    curso_id = request.POST.get('curso_id')
    profesor_id = request.POST.get('profesor_id')
    cursito = get_object_or_404(curso, pk=curso_id)
    profesor = get_object_or_404(Estudiante, pk=profesor_id, tipo_estudiante='profesor')
    cursito.estudiante = profesor
    cursito.save()
    return JsonResponse({'ok': True})

@login_required
@estudiante_tipo_requerido(['administrador'])
@require_POST
def desasignar_curso(request):
    curso_id = request.POST.get('curso_id')
    cursito = get_object_or_404(curso, pk=curso_id)
    cursito.estudiante = None
    cursito.save()
    return JsonResponse({'ok': True})






#################3
@login_required
def vista_grafico_estudiante(request):
    base_template = "layouts/base2.html"
    usuario = request.user
    return render(request, 'grafico_estudiante.html', {'usuario': usuario, 'base_template': base_template})


@login_required
@estudiante_tipo_requerido(['administrador'])
def vista_grafico_respuestas(request):
    base_template = "layouts/base.html"
    estudiantes = Estudiante.objects.filter(tipo_estudiante='estudiante')
    return render(request, 'grafico_respuestas.html', {'estudiantes': estudiantes, 'base_template': base_template})


@login_required
@estudiante_tipo_requerido(['tutor'])
def vista_grafico_respuestas_tutor(request):
    usuario_actual = request.user
    base_template = "layouts/base_tutor.html"
    estudiantes = Estudiante.objects.filter(tipo_estudiante='estudiante', tutor_id = usuario_actual)
    return render(request, 'grafico_respuestas.html', {'estudiantes': estudiantes, 'base_template': base_template})

def obtener_reportes_resumen(request):
    usuario = request.GET.get('usuario')
    curso = request.GET.get('curso')

    if not usuario or not curso:
        return JsonResponse({'error': 'Datos incompletos'}, status=400)

    estudiante = Estudiante.objects.get(usuario=usuario)
    reportes = Reporte.objects.filter(KK_usuario=estudiante).order_by('fecha_de_examen')
    variable = VariableControl.objects.first()

    datos = []
    for r in reportes:
        correctas = getattr(r, f'{curso}_1')
        incorrectas = getattr(r, f'{curso}_2')
        nivel = r.nivel

        if nivel == 90:  # Preuniversitario
            total_preguntas = getattr(variable, f'{curso}_Pre')
        else:
            total_preguntas = getattr(variable, f'{curso}_Sem')

        blancas = total_preguntas - (correctas + incorrectas)
        datos.append({
            'fecha': r.fecha_de_examen.strftime('%d/%m/%Y'),
            'correctas': correctas,
            'incorrectas': incorrectas,
            'blancas': max(blancas, 0)
        })

    return JsonResponse(datos, safe=False)




@estudiante_tipo_requerido(['administrador'])
def iniciar_tarea_reportes(request):
    task = generar_reportes_completos_task.delay()
    return JsonResponse({'task_id': task.id})


def obtener_estado_tarea(request, task_id):
    result = AsyncResult(task_id)
    data = {
        'state': result.state,
        'info': result.info if result.info else {}
    }
    return JsonResponse(data)



@login_required
@estudiante_tipo_requerido(['tutor'])
def estudiantes_asignados_tutor(request):
    base_template = "layouts/base_tutor.html"
    tutor = request.user
    estudiantes = Estudiante.objects.filter(tutor_id=tutor)

    return render(request, 'estudiantes_asignados.html', {
        'base_template': base_template,
        'tutor': tutor,
        'estudiantes': estudiantes,
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def listar_tutores(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    tutores = Estudiante.objects.filter(tipo_estudiante='tutor')
    return render(request, 'listar_tutores.html', {
        'tutores': tutores,
        'base_template': base_template,
    })

def obtener_todos_los_estudiantes(request):
    estudiantes = Estudiante.objects.filter(tipo_estudiante = 'estudiante').select_related('tutor')
    data = [
        {
            'ID_Estudiante': est.ID_Estudiante,
            'usuario': est.usuario,
            'nombre': est.nombre,
            'tutor': est.tutor.nombre if est.tutor else None,
        }
        for est in estudiantes
    ]
    return JsonResponse(data, safe=False)



@login_required
@estudiante_tipo_requerido(['administrador'])
def crear_tutor(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        usuario = request.POST.get('usuario')
        contraseña = request.POST.get('contraseña')

        if Estudiante.objects.filter(usuario=usuario).exists():
            return render(request, 'crear_tutor.html', {
                'mensaje': 'El nombre de usuario ya existe.'
            })

        Estudiante.objects.create(
            nombre=nombre,
            usuario=usuario,
            contraseña=contraseña,
            tipo_estudiante='tutor',
            is_active=True,
            is_staff=False,
        )
        return redirect('listar_tutores')

    return render(request, 'crear_tutor.html', {
        'base_template': base_template,
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def asignar_tutores(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"

    tutores = Estudiante.objects.filter(tipo_estudiante='tutor')
    estudiantes = Estudiante.objects.filter(tipo_estudiante='estudiante')

    return render(request, 'asignar_tutores.html', {
        'base_template': base_template,
        'tutores': tutores,
        'estudiantes': estudiantes,
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def obtener_estudiantes_tutor(request, tutor_id):
    tutor = get_object_or_404(Estudiante, pk=tutor_id)
    estudiantes = tutor.tutorados.all().values('ID_Estudiante', 'nombre', 'usuario')
    return JsonResponse(list(estudiantes), safe=False)


@login_required
@estudiante_tipo_requerido(['administrador', 'tutor'])
def reportes_estudiante_tutor(request, pk):

    usuario = request.user

    usuario_actual = get_object_or_404(Estudiante, pk=pk)
    
    if usuario.tipo_estudiante == "administrador":
        base_template = "layouts/base.html"
    elif usuario.tipo_estudiante == "tutor":
        base_template = "layouts/base_tutor.html"
    else:
        base_template = "layouts/base2.html"

    reportes = Reporte.objects.filter(KK_usuario=usuario_actual).order_by('-fecha_de_examen')
    print(reportes)
    cursos = [
        "Razonamiento Verbal", "Razonamiento Matemático", "Aritmética", "Álgebra",
        "Geometría", "Trigonometría", "Física", "Química", "Biología", "Lenguaje",
        "Literatura", "Historia", "Geografía", "Filosofía", "Psicología", "Economía"
    ]

    fechas = list(reportes.values_list('fecha_de_examen', flat=True).distinct())
    graficos = []

    # Ruta base dentro de MEDIA_ROOT
    carpeta_usuario = os.path.join(settings.MEDIA_ROOT, f"{usuario_actual.usuario}_2025")
    print(usuario_actual)
    for reporte in reportes:
        datos = reporte.obtener_datos()

        for curso in datos.keys():
            # Generar el nombre del archivo según la estructura "curso_usuario_fecha.png"
            nombre_archivo = f"{curso}_{usuario_actual.usuario}_{reporte.fecha_de_examen}.png".replace(" ", "_")
            ruta_relativa = os.path.join(f"{usuario_actual.usuario}_2025", nombre_archivo)
            ruta_absoluta = os.path.join(settings.MEDIA_ROOT, ruta_relativa)

            # Verificar si el archivo realmente existe antes de agregarlo
            if os.path.exists(ruta_absoluta):
                graficos.append({
                    "fecha": reporte.fecha_de_examen,
                    "curso": curso,
                    "ruta_imagen": f"{settings.MEDIA_URL}{ruta_relativa.replace('\\', '/')}"
                })
    
    return render(request, 'home.html', {
        'graficos': graficos,
        'cursos': cursos,
        'fechas': fechas,
        'base_template': base_template,
        'usuario_actual': usuario_actual,  # <--- esto
    })

@login_required
@estudiante_tipo_requerido(['administrador', 'tutor'])
def ver_asistencias_tutor(request, pk):
    usuario = request.user
    estudiante = get_object_or_404(Estudiante, pk=pk)
    if usuario.tipo_estudiante == "administrador":
        base_template = "layouts/base.html"
    elif usuario.tipo_estudiante == "tutor":
        base_template = "layouts/base_tutor.html"
    else:
        base_template = "layouts/base2.html"
    fecha_filtrada = request.GET.get('fecha')
    # Filtrar asistencias del estudiante y ordenarlas
    asistencias = Asistencia.objects.filter(KK_usuario=estudiante).order_by('Fecha', 'Hora')

    if fecha_filtrada:
        try:
            fecha_obj = datetime.strptime(fecha_filtrada, "%Y-%m-%d").date()
            asistencias = asistencias.filter(Fecha=fecha_obj)
        except ValueError:
            pass  # Fecha inválida, ignora el filtro

    agrupadas_por_fecha = defaultdict(list)
    for asistencia in asistencias:
        agrupadas_por_fecha[asistencia.Fecha].append(asistencia)

    datos = []
    for fecha in asistencias.values_list('Fecha', flat=True).distinct().order_by('Fecha'):
        marcas = asistencias.filter(Fecha=fecha).order_by('Hora')
        for i, asistencia in enumerate(marcas, start=1):
            dia_semana = dias_semana[fecha.weekday()]
            fecha_corta = f"{dia_semana} {fecha.strftime('%d/%m')}"
            datos.append({
                "usuario": estudiante.usuario,
                "nombre": estudiante.nombre,
                "numero_marca": i,
                "fecha": fecha.strftime('%d/%m/%Y'),
                "fecha_corta": fecha_corta,
                "hora": asistencia.Hora,
                "observacion": asistencia.Observacion,
                "modalidad" : asistencia.Modalidad,
            })

    return render(request, "ver_asistencias.html", {"asistencias": datos, "base_template": base_template, "fecha_filtrada": fecha_filtrada})


@login_required
@estudiante_tipo_requerido(['administrador'])
def asignar_estudiante(request):
    if request.method == 'POST':
        estudiante_id = request.POST.get('estudiante_id')
        tutor_id = request.POST.get('tutor_id')
        estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
        tutor = get_object_or_404(Estudiante, pk=tutor_id)
        estudiante.tutor = tutor
        estudiante.save()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
@estudiante_tipo_requerido(['administrador'])
def desasignar_estudiante(request):
    if request.method == 'POST':
        estudiante_id = request.POST.get('estudiante_id')
        estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
        estudiante.tutor = None
        estudiante.save()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


#Descargas

@login_required
@estudiante_tipo_requerido(['administrador'])
@require_POST
@csrf_exempt
def descargar_estudiantes_excel(request):
    ids = request.POST.getlist('estudiantes')
    if not ids:
        return HttpResponse("No se seleccionaron estudiantes", status=400)

    estudiantes = Estudiante.objects.filter(pk__in=ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    # Cabeceras
    headers = ['Codigo', 'Nombre', 'Facebook', 'Instagram', 'Celular', 'Colegio', 'Grado', 'Ciudad', 'Carrera']
    ws.append(headers)

    # Datos
    for est in estudiantes:
        ws.append([
            est.usuario,
            est.nombre,
            est.facebook,
            est.instagram,
            est.numCelular,
            est.colegio,
            est.grado,
            est.ciudad,
            est.carrera,
        ])

    # Ajustar ancho de columnas
    for i, col in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=estudiantes.xlsx'
    wb.save(response)
    return response


@login_required
def descargar_reporte(request, pk):
    # Buscar al estudiante por su PK
    estudiante = get_object_or_404(Estudiante, pk=pk)

    try:
        # Ruta donde debería estar el reporte del estudiante
        ruta_descarga = os.path.join("media", "reportes", "simulacros", f"{estudiante.usuario}_reporte_simulacro.png")
        if not os.path.exists(ruta_descarga):
            raise Http404("El reporte no se ha generado todavía, por favor ponte en contacto con la administración.")
        
        return FileResponse(open(ruta_descarga, 'rb'), as_attachment=True, filename=os.path.basename(ruta_descarga))
    except Exception as e:
        print(f"Error al descargar el reporte: {e}")
        raise Http404("El reporte no se ha generado todavía, por favor ponte en contacto con la administración.")



@require_POST
@login_required
@estudiante_tipo_requerido(['administrador'])
def descargar_reportes_zip(request):
    # Obtener los IDs enviados por el formulario
    ids = request.POST.getlist('estudiantes')
    ids = [id for id in ids if id.isdigit()]
    
    if not ids:
        messages.warning(request, "No seleccionaste ningún estudiante válido.")
        return redirect('lista_estudiantes')

    estudiantes = Estudiante.objects.filter(ID_Estudiante__in=ids)
    carpeta = os.path.join(settings.MEDIA_ROOT, "reportes", "simulacros")
    
    if not os.path.exists(carpeta):
        raise Http404("La carpeta de reportes no existe")

    # Crear ZIP en memoria
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for estudiante in estudiantes:
            nombre_archivo = f"{estudiante.usuario}_reporte_simulacro.png"
            ruta_archivo = os.path.join(carpeta, nombre_archivo)
            if os.path.isfile(ruta_archivo):
                zip_file.write(ruta_archivo, arcname=nombre_archivo)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="reportes_seleccionados.zip"'
    return response


@require_POST
@login_required
@estudiante_tipo_requerido(['administrador'])
def descargar_reportes_asistencia_zip(request):
    ids = request.POST.getlist('estudiantes')
    ids = [id for id in ids if id.isdigit()]

    if not ids:
        raise Http404("No se seleccionaron estudiantes válidos.")

    estudiantes = Estudiante.objects.filter(ID_Estudiante__in=ids)

    carpeta = os.path.join(settings.MEDIA_ROOT, "reportes", "asistencias")
    if not os.path.exists(carpeta):
        raise Http404("La carpeta de reportes no existe, comuníquese con soporte")

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for estudiante in estudiantes:
            archivo = os.path.join(carpeta, f"{estudiante.usuario}.png")
            if os.path.isfile(archivo):
                zip_file.write(archivo, arcname=os.path.basename(archivo))

    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="reportes_asistencias.zip"'
    return response


#Generación/cración de imágenes/reportes

def generar_reporte_asistencia_todos(request):
    try:
        estudiantes = Estudiante.objects.all()
        for estudiante in estudiantes:
            asistencia = estudiante.asistencia.all().order_by('Fecha')
            generar_reporte_asistencia_imagen(estudiante, asistencia)
        messages.success(request, "Gráficos y reportes generados exitosamente para todos los estudiantes.")
        return redirect('seleccionar_fecha_generacion')
    except Exception as e:
        print(f"Error al generar todo el reporte: {e}")
        raise Http404("Ocurrió un error al generar los reportes.{e}")




@login_required
@estudiante_tipo_requerido(['administrador'])
def editar_variable_control(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    variable = get_object_or_404(VariableControl, pk=1)
    if request.method == 'POST':
        form = VariableControlForm(request.POST, instance=variable)
        if form.is_valid():
            posiciones = [
                (556, 655), (1320, 655), (556, 893), (1320, 893), (556, 1133), (1320, 1131), (556, 1370), (1320, 1370),
                (556, 1612), (1320, 1614), (556, 1853), (1320, 1853), (616, 2097), (1365, 2097), (556, 2333), (1320, 2334)
            ]

            def generar_plantilla(cantidades, output_filename):
                plantilla_path = os.path.join(settings.BASE_DIR, "login/static/img/plantilla-plantilla-notas.png")
                #plantilla_path = os.path.join(settings.BASE_DIR, 'login', 'static', 'img', 'plantilla-plantilla-notas.png')
                salida_path = os.path.join(settings.BASE_DIR, "login/static/img/", output_filename)
                
                imagen = Image.open(plantilla_path)
                try:
                    fuente = ImageFont.truetype("arial.ttf", 22)
                except:
                    fuente = ImageFont.load_default()

                draw = ImageDraw.Draw(imagen)
                for valor, posicion in zip(cantidades, posiciones):
                    draw.text(posicion, str(valor), fill="white", font=fuente)

                imagen.save(salida_path)

            # Obtener instancia guardada de VariableControl
            variable_control = form.save(commit=False)

            # Obtener cantidades
            Cantidad_Preguntas_Pre = [
                variable_control.Ar_Pre, variable_control.Bi_Pre, variable_control.Ec_Pre, variable_control.Fil_Pre,
                variable_control.Fi_Pre, variable_control.Gf_Pre, variable_control.Ge_Pre, variable_control.Hi_Pre,
                variable_control.Le_Pre, variable_control.Lit_Pre, variable_control.Psi_Pre, variable_control.Qu_Pre,
                variable_control.Rm_Pre, variable_control.Rv_Pre, variable_control.Tr_Pre, variable_control.Al_Pre
            ]

            Cantidad_Preguntas_Sem = [
                variable_control.Ar_Sem, variable_control.Bi_Sem, variable_control.Ec_Sem, variable_control.Fil_Sem,
                variable_control.Fi_Sem, variable_control.Gf_Sem, variable_control.Ge_Sem, variable_control.Hi_Sem,
                variable_control.Le_Sem, variable_control.Lit_Sem, variable_control.Psi_Sem, variable_control.Qu_Sem,
                variable_control.Rm_Sem, variable_control.Rv_Sem, variable_control.Tr_Sem, variable_control.Al_Sem
            ]
            # Generar imágenes
            generar_plantilla(Cantidad_Preguntas_Pre, 'plantilla-notas-pre.png')
            generar_plantilla(Cantidad_Preguntas_Sem, 'plantilla-notas-semillas.png')

            # Guardar finalmente el formulario
            variable_control.save()
            form.save()
            messages.success(request, "Valores y plantillas actualizadas")
            return redirect('editar_variable_control')
    else:
        form = VariableControlForm(instance=variable)

    # Diccionario de nombres personalizados
    field_labels = {
        'Rm_Pre': 'Raz. Matemático', 'Rv_Pre': 'Raz. Verbal', 'Ar_Pre': 'Aritmética',
        'Al_Pre': 'Álgebra', 'Ge_Pre': 'Geometría', 'Tr_Pre': 'Trigonometría',
        'Fi_Pre': 'Física', 'Qu_Pre': 'Química', 'Bi_Pre': 'Biología',
        'Le_Pre': 'Lenguaje', 'Lit_Pre': 'Literatura', 'Hi_Pre': 'Historia',
        'Gf_Pre': 'Geografía', 'Fil_Pre': 'Filosofía', 'Psi_Pre': 'Psicología',
        'Ec_Pre': 'Economía',

        'Rm_Sem': 'Raz. Matemático', 'Rv_Sem': 'Raz. Verbal', 'Ar_Sem': 'Aritmética',
        'Al_Sem': 'Álgebra', 'Ge_Sem': 'Geometría', 'Tr_Sem': 'Trigonometría',
        'Fi_Sem': 'Física', 'Qu_Sem': 'Química', 'Bi_Sem': 'Biología',
        'Le_Sem': 'Lenguaje', 'Lit_Sem': 'Literatura', 'Hi_Sem': 'Historia',
        'Gf_Sem': 'Geografía', 'Fil_Sem': 'Filosofía', 'Psi_Sem': 'Psicología',
        'Ec_Sem': 'Economía',

        'EntradaManana': 'Entrada Mañana', 'SalidaManana': 'Salida Mañana',
        'EntradaTarde': 'Entrada Tarde', 'SalidaTarde': 'Salida Tarde',
        'EntradaAmanecida': 'Entrada Amanecida', 'SalidaAmanecida': 'Salida Amanecida',

        "Preg_Corr_Raz": 'Pregunta Correcta Raz.', "Preg_Corr_Con": 'Pregunta Correcta Con.',
        "Preg_Inc_Raz": 'Pregunta Incorrecta Raz.', "Preg_Inc_Con": 'Pregunta Incorrecta Con.'
    }

    preguntas_pre = [
        'Rm_Pre', 'Rv_Pre', 'Ar_Pre', 'Al_Pre', 'Ge_Pre', 'Tr_Pre',
        'Fi_Pre', 'Qu_Pre', 'Bi_Pre', 'Le_Pre', 'Lit_Pre', 'Hi_Pre',
        'Gf_Pre', 'Fil_Pre', 'Psi_Pre', 'Ec_Pre'
    ]
    preguntas_sem = [
        'Rm_Sem', 'Rv_Sem', 'Ar_Sem', 'Al_Sem', 'Ge_Sem', 'Tr_Sem',
        'Fi_Sem', 'Qu_Sem', 'Bi_Sem', 'Le_Sem', 'Lit_Sem', 'Hi_Sem',
        'Gf_Sem', 'Fil_Sem', 'Psi_Sem', 'Ec_Sem'
    ]
    horarios = [
        'EntradaManana', 'SalidaManana', 'EntradaTarde',
        'SalidaTarde', 'EntradaAmanecida', 'SalidaAmanecida'
    ]

    puntajes = [
        'Preg_Corr_Raz', 'Preg_Corr_Con', 'Preg_Inc_Raz',
        'Preg_Inc_Con'
    ]

    return render(request, 'editar_variable_control.html', {
        'base_template': base_template,
        'form': form,
        'preguntas_pre': preguntas_pre,
        'preguntas_sem': preguntas_sem,
        'horarios': horarios,
        'puntajes': puntajes,
        'field_labels': field_labels,
    })


def generar_reporte_asistencia_imagen(estudiante, asistencias):
    
    plantilla_path = os.path.join(settings.MEDIA_ROOT, "login/static/img/plantilla-asistencia.png")  # Asegúrate de que esta plantilla exista
    imagen = Image.open(plantilla_path)
    draw = ImageDraw.Draw(imagen)

    try:
        fuente_titulo = ImageFont.truetype("arial.ttf", 28)
        fuente_texto = ImageFont.truetype("arial.ttf", 22)
    except:
        fuente_titulo = ImageFont.load_default()
        fuente_texto = ImageFont.load_default()

    posiciones_x = {
        "usuario": 88,
        "nombre": 185,
        "Fecha_Asistencia": 602,
        "Ingreso_mana": 740,
        "Salida_mana": 820,
        "Ingreso_tarde": 915,
        "Salida_tarde": 995,
        "Ingreso_noche": 1085,
        "Salida_noche": 1165,
        "Observacion": 1250,
    }

    b = 300
    for asistencia in asistencias:
        fila = {
            "usuario": estudiante.usuario,
            "nombre": estudiante.nombre,
            "Fecha_Asistencia": asistencia.Fecha.strftime('%d/%m/%Y'),
            "Ingreso_mana": asistencia.Ingreso_mana,
            "Salida_mana": asistencia.Salida_mana,
            "Ingreso_tarde": asistencia.Ingreso_tarde,
            "Salida_tarde": asistencia.Salida_tarde,
            "Ingreso_noche": asistencia.Ingreso_noche,
            "Salida_noche": asistencia.Salida_noche,
            "Observacion": asistencia.Observacion,
        }
        for campo, x in posiciones_x.items():
            draw.text((x, b), str(fila[campo]), fill="black", font=fuente_texto)
        b += 45

    carpeta = f'media/reportes/asistencias'
    os.makedirs(carpeta, exist_ok=True)
    nombre_archivo = os.path.join(carpeta, f'{estudiante.usuario}_reporte_asistencia.png')
    imagen.save(nombre_archivo)

    return nombre_archivo

@login_required
def generar_reporte_asistencia(request):
    estudiante = request.user  # Asume que Estudiante es el modelo de usuario
    asistencias = Asistencia.objects.filter(KK_usuario=estudiante).order_by('Fecha')
    if not asistencias.exists():
        return render(request, "prueba.html")  # Puedes crear esta plantilla si deseas

    ruta_imagen = generar_reporte_asistencia_imagen(estudiante, asistencias)

    return FileResponse(open(ruta_imagen, 'rb'), as_attachment=True, filename=os.path.basename(ruta_imagen))


cursos_pre = [
            "Razonamiento Verbal", "Razonamiento Matemático", "Aritmética", "Álgebra",
            "Geometría", "Trigonometría", "Física", "Química", "Biología", "Lenguaje",
            "Literatura", "Historia", "Geografía", "Filosofía", "Psicología", "Economía"
        ]

preguntas_pre = {
            "Razonamiento Verbal": 15,
            "Razonamiento Matemático": 15,
            "Aritmética": 5,
            "Álgebra": 5,
            "Geometría": 4,
            "Trigonometría": 4,
            "Física": 5,
            "Química": 5,
            "Biología": 9,
            "Lenguaje": 6,
            "Literatura": 6,
            "Historia": 3,
            "Geografía": 2,
            "Filosofía": 2,
            "Psicología": 2,
            "Economía": 2,
        }


preguntas_semillero = {
            "Razonamiento Verbal": 10,
            "Razonamiento Matemático": 10,
            "Aritmética": 5,
            "Álgebra": 5,
            "Geometría": 0,
            "Trigonometría": 0,
            "Física": 0,
            "Química": 0,
            "Biología": 0,
            "Lenguaje": 0,
            "Literatura": 5,
            "Historia": 0,
            "Geografía": 0,
            "Filosofía": 0,
            "Psicología": 0,
            "Economía": 0,
        }


@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_todos_los_reportes(request):
    generar_reportes_completos_task.delay()
    messages.success(request, "Se están generando todos los reportes en segundo plano. Puedes continuar usando el sistema.")
    return redirect('seleccionar_fecha_generacion')

@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_todo_reporte(request):
    # Lanzar la tarea maestra que engloba la generación de reportes e imágenes
    generar_reportes_completos_task.delay()
    messages.success(request, "Se está generando el reporte completo en segundo plano. Puedes continuar usando el sistema.")
    return redirect('seleccionar_fecha_generacion')

def crear_grafico_estudiante_curso(estudiante, nombre_curso, datos):
    if not datos:
        return

    df = pd.DataFrame(datos)
    if df.empty:
        return

    df['Fecha Simulacro'] = pd.to_datetime(df['Fecha Simulacro'])
    df = df.sort_values('Fecha Simulacro')

    df['blanco'] = df['total'] - df['correctas'] - df['incorrectas']

    fechas = df['Fecha Simulacro'].unique()
    cols = 4
    rows = math.ceil(len(fechas) / cols)

    fig, axes = plt.subplots(rows, cols, figsize=(20, 5 * rows))
    axes = axes.flatten()

    for i, fecha in enumerate(fechas):
        datos_fecha = df[df['Fecha Simulacro'] == fecha]
        ax = axes[i]

        index = range(len(datos_fecha))
        bar_width = 0.2

        bars1 = ax.bar(index, datos_fecha['correctas'] * 0.85, bar_width, color='green')
        bars2 = ax.bar([i + bar_width for i in index], datos_fecha['incorrectas'] * 0.85, bar_width, color='red')
        bars3 = ax.bar([i + 2 * bar_width for i in index], datos_fecha['blanco'] * 0.85, bar_width, color='gray')

        for spine in ax.spines.values():
            spine.set_visible(False)

        ax.set_xticks([])
        ax.set_xlabel(fecha.strftime('%d/%m/%Y'), fontsize=30)
        ax.set_yticks([])

        for bars, original_values in zip([bars1, bars2, bars3], [datos_fecha['correctas'], datos_fecha['incorrectas'], datos_fecha['blanco']]):
            for bar, original_value in zip(bars, original_values):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, height + 0.03, f"{int(original_value)}", ha='center', fontsize=40, fontweight='bold')

    for j in range(i + 1, len(axes)):
        fig.delaxes(axes[j])

    plt.subplots_adjust(bottom=0.15)

    carpeta = f'media/{estudiante.usuario}_2025/imagenes_reporte/'
    os.makedirs(carpeta, exist_ok=True)
    img_path = os.path.join(carpeta, f"{nombre_curso.replace(' ', '_')}.png")
    plt.savefig(img_path, dpi=300)
    plt.close()



@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_graficos_todos_estudiantes(request):
    estudiantes = Estudiante.objects.filter(reporte_actualizado = True)

    for estudiante in estudiantes:
        reportes = estudiante.reportes.all().order_by('fecha_de_examen')
        if not reportes.exists():
            continue

        # Diccionario con listas de dicts por curso
        cursos_data = {}

        for reporte in reportes:
            fecha = reporte.fecha_de_examen
            datos_cursos = reporte.obtener_datos()  # ✅ Aquí obtenemos (correctas, incorrectas)

            for curso, (corr, inc) in datos_cursos.items():
                if curso not in cursos_data:
                    cursos_data[curso] = []

                cursos_data[curso].append({
                    'Fecha Simulacro': fecha,
                    'correctas': corr,
                    'incorrectas': inc,
                })

        # Crear un gráfico por curso
        for curso, datos in cursos_data.items():
            crear_grafico_estudiante_curso(estudiante, curso, datos)

    messages.success(request, "Gráficos generados exitosamente para todos los estudiantes.")
    return redirect('seleccionar_fecha_generacion')

@login_required
@estudiante_tipo_requerido(['administrador'])
def seleccionar_fecha_generacion(request):
    fechas_disponibles = Reporte.objects.values_list('fecha_de_examen', flat=True).distinct().order_by('-fecha_de_examen')

    if request.method == "POST":
        fecha_str = request.POST.get("fecha")
        return redirect('generar_imagenes_fecha', fecha=fecha_str)

    return render(request, 'seleccionar_fecha_reporte.html', {
        'fechas': fechas_disponibles,
        "base_template": "layouts/base.html"
    })

@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_imagenes_reportes_por_fecha(request):
    # Lanzar la tarea maestra que incluye la generación de imágenes
    generar_reportes_completos_task.delay()
    messages.success(request, "Se está generando el reporte completo (incluye imágenes) en segundo plano.")
    return redirect('seleccionar_fecha_generacion')

@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_reportes_odf_asistencia(request):
    asistencias = Asistencia.objects.all()
    asistencias_ids = list(Asistencia.objects.values_list('ID_Reporte', flat=True))
    generar_pdfs_asistencias_por_estudiante_task.delay(asistencias_ids)
    messages.success(request, "Se están generando los reportes para las asistencias en segundo plano. Puedes continuar usando el sistema.")
    return redirect('seleccionar_fecha_generacion')

@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_reportes_odf_asistencia_3(request):
    asistencias = Asistencia.objects.all()
    generar_reportes_odf_asistencia_1(asistencias)
    messages.success(request, "Se están generando los reportes para las asistencias en segundo plano. Puedes continuar usando el sistema.")
    return redirect('seleccionar_fecha_generacion')

def generar_reportes_odf_asistencia_1(asistencias_queryset):
    #asistencias_queryset = Asistencia.objects.all()
    carpeta = os.path.join(settings.MEDIA_ROOT, "reportes", "asistencias")
    os.makedirs(carpeta, exist_ok=True)

    # Agrupamos asistencias por usuario
    asistencias_por_usuario = defaultdict(list)
    for a in asistencias_queryset.order_by('KK_usuario__usuario', 'Fecha', 'Hora'):
        asistencias_por_usuario[a.KK_usuario.usuario].append(a)

    for usuario, asistencias_usuario in asistencias_por_usuario.items():
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)

        def dibujar_encabezado(y_pos):
            p.setFont("Helvetica-Bold", 14)
            p.drawString(30, height - 40, f"Reporte de Asistencias - {usuario}")

            headers = ["Usuario", "Nombre", "Fecha", "Hora", "Nro Marca", "Modalidad", "Observación"]
            x_positions = [30, 80, 310, 390, 470, 550, 650]
            p.setFont("Helvetica-Bold", 10)
            for i, header in enumerate(headers):
                p.drawString(x_positions[i], y_pos, header)
            return y_pos - 25

        y = dibujar_encabezado(height - 80)
        p.setFont("Helvetica", 10)

        # Agrupamos por (usuario, fecha) para asignar Nro Marca
        agrupadas = defaultdict(list)
        for a in asistencias_usuario:
            clave = (a.KK_usuario.usuario, a.Fecha)
            agrupadas[clave].append(a)

        for grupo in agrupadas.values():
            for i, a in enumerate(grupo, start=1):
                datos = [
                    a.KK_usuario.usuario,
                    a.KK_usuario.nombre,
                    a.Fecha.strftime('%d/%m/%Y'),
                    a.Hora,
                    "ENTRADA" if i in [1, 3] else "SALIDA",
                    a.Modalidad,
                    a.Observacion or "-"
                ]
                x_positions = [30, 80, 310, 390, 470, 550, 650]
                for j, valor in enumerate(datos):
                    p.drawString(x_positions[j], y, str(valor))
                y -= 20

                if y < 40:
                    p.showPage()
                    y = dibujar_encabezado(height - 50)
                    p.setFont("Helvetica", 10)

        p.save()
        buffer.seek(0)

        # Guardamos el archivo
        nombre_archivo = f"{usuario}_reporte_asistencia.pdf"
        ruta_archivo = os.path.join(carpeta, nombre_archivo)
        with open(ruta_archivo, 'wb') as f:
            f.write(buffer.getbuffer())

def generar_imagenes_reportes_por_fecha_respaldo(request, fecha):
    fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    cursos = [
        "Razonamiento Verbal", "Razonamiento Matemático", "Aritmética", "Álgebra",
        "Geometría", "Trigonometría", "Física", "Química", "Biología", "Lenguaje",
        "Literatura", "Historia", "Geografía", "Filosofía", "Psicología", "Economía"
    ]

    ruta_base = 'media'

    for estudiante in Estudiante.objects.filter(reporte_actualizado = True):
        reportes = Reporte.objects.filter(KK_usuario=estudiante, fecha_de_examen=fecha_obj)

        for reporte in reportes:
            datos = reporte.obtener_datos()
            nivel = reporte.nivel  # 30 para semillero, 40 para pre u otro valor que determines

            if nivel == 30:
                preguntas_por_curso = preguntas_semillero
            else:
                preguntas_por_curso = preguntas_pre

            carpeta_usuario = os.path.join(ruta_base, f"{estudiante.usuario}_2025")
            os.makedirs(carpeta_usuario, exist_ok=True)

            for curso, valores in datos.items():
                correctas, incorrectas = valores
                total_preguntas = preguntas_por_curso.get(curso, 10)
                en_blanco = total_preguntas - (correctas + incorrectas)

                fig, ax = plt.subplots(figsize=(8, 4))
                ax.bar(["Correctas", "Incorrectas", "En blanco"],
                       [correctas, incorrectas, en_blanco],
                       color=["green", "red", "gray"])
                ax.set_title(f"{curso} - {reporte.fecha_de_examen}")

                nombre_archivo = f"{curso}_{estudiante.usuario}_{reporte.fecha_de_examen}.png".replace(" ", "_")
                ruta_archivo = os.path.join(carpeta_usuario, nombre_archivo)

                plt.savefig(ruta_archivo, format='png')
                plt.close(fig)

    return redirect('seleccionar_fecha_generacion')

@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_reporte(request):
    usuario_actual = request.user
    try:
        estudiantes = Estudiante.objects.filter(reporte_actualizado = True)
        ruta_guardar = f'media/reportes/simulacros'
        plantilla_path = os.path.join(settings.MEDIA_ROOT, "login/static/img/plantilla-notas.png")

        for estudiante in estudiantes:
            ruta_carpeta = f'media/{estudiante.usuario}_2025/imagenes_reporte/'
            imagenes = [
                "Aritmética.png", "Biología.png", "Economía.png", "Filosofía.png",
                "Física.png", "Geografía.png", "Geometría.png", "Historia.png",
                "Lenguaje.png", "Literatura.png", "Psicología.png", "Química.png",
                "Razonamiento_Matemático.png", "Razonamiento_Verbal.png", "Álgebra.png",
                "Trigonometría.png",
            ]

            # Cargar la plantilla
            # Asegúrate de que esta plantilla exista
            plantilla = Image.open(plantilla_path)

            # Crear un objeto de dibujo
            draw = ImageDraw.Draw(plantilla)
            fuente = ImageFont.truetype("arial.ttf", 40)
            draw.text((720, 124), estudiante.nombre, fill="black", font=fuente)

            # Añadir imágenes
            a = 180
            b = 210
            i = 1
            for imagen in imagenes:
                ruta_imagen = os.path.join(ruta_carpeta, imagen)
                if not os.path.exists(ruta_imagen):
                    continue
                if i % 2 == 0:
                    a = 960
                else:
                    a = 180
                    b += 240
                otra_imagen = Image.open(ruta_imagen)
                imagen_redimensionada = otra_imagen.resize((700, 200))
                plantilla.paste(imagen_redimensionada, (a, b))
                i += 1

            resultado_path = os.path.join(ruta_guardar, f"{estudiante.usuario}_reporte_simulacro.png")
            plantilla.save(resultado_path)

            # 🔽 Aquí está el cambio importante
        messages.success(request, "Reportes generados exitosamente para todos los estudiantes.")
        return redirect("seleccionar_fecha_generacion")

    except Exception as e:
        print(f"Error al generar el reporte: {e}")
        raise Http404("No se pudo generar el reporte{e}")

@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_reportes_simulacro_todos(request):
    try:
        estudiantes = Estudiante.objects.filter(reporte_actualizado = True)
        plantilla_path = os.path.join(settings.BASE_DIR, "login/static/img/plantilla-notas.png")
        fuente_path = os.path.join(settings.BASE_DIR, "static/fonts/arial.ttf")  # Asegúrate que exista

        for estudiante in estudiantes:
            ruta_carpeta = f'media/{estudiante.usuario}_2025/imagenes_reporte/'
            imagenes = [
                "Aritmética.png", "Biología.png", "Economía.png", "Filosofía.png",
                "Física.png", "Geografía.png", "Geometría.png", "Historia.png",
                "Lenguaje.png", "Literatura.png", "Psicología.png", "Química.png",
                "Razonamiento_Matemático.png", "Razonamiento_Verbal.png", "Álgebra.png",
                "Trigonometría.png",
            ]

            if not os.path.exists(plantilla_path):
                continue

            plantilla = Image.open(plantilla_path)
            draw = ImageDraw.Draw(plantilla)

            try:
                fuente = ImageFont.truetype(fuente_path, 40)
            except:
                fuente = ImageFont.load_default()

            draw.text((720, 124), estudiante.nombre, fill="black", font=fuente)

            a = 180
            b = 210
            i = 1

            for imagen in imagenes:
                ruta_imagen = os.path.join(ruta_carpeta, imagen)
                if not os.path.exists(ruta_imagen):
                    continue

                if i % 2 == 0:
                    a = 960
                else:
                    a = 180
                    b += 240

                otra_imagen = Image.open(ruta_imagen)
                imagen_redimensionada = otra_imagen.resize((700, 200))
                plantilla.paste(imagen_redimensionada, (a, b))
                i += 1

            # Guardar en carpeta final
            carpeta_destino = f'media/reportes/simulacros/'
            os.makedirs(carpeta_destino, exist_ok=True)
            resultado_path = os.path.join(carpeta_destino, f"{estudiante.usuario}_reporte_simulacro.png")
            plantilla.save(resultado_path)

        messages.success(request, "Reportes generados exitosamente para todos los estudiantes.")
        return redirect("seleccionar_fecha_generacion")

    except Exception as e:
        print(f"Error al generar reportes: {e}")
        raise Http404("Error al generar los reportes.{e}")

@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_imagenes_reportes(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"

    cursos_pre = [
        "Razonamiento Verbal", "Razonamiento Matemático", "Aritmética", "Álgebra",
        "Geometría", "Trigonometría", "Física", "Química", "Biología", "Lenguaje",
        "Literatura", "Historia", "Geografía", "Filosofía", "Psicología", "Economía"
    ]

    cursos_semillero = [
        "Razonamiento Verbal", "Razonamiento Matemático", "Aritmética", "Álgebra", "Literatura"
    ]

    preguntas_por_curso_pre = {
        "Razonamiento Verbal": 15,
        "Razonamiento Matemático": 15,
        "Aritmética": 5,
        "Álgebra": 5,
        "Geometría": 4,
        "Trigonometría": 4,
        "Física": 5,
        "Química": 5,
        "Biología": 9,
        "Lenguaje": 6,
        "Literatura": 6,
        "Historia": 3,
        "Geografía": 2,
        "Filosofía": 2,
        "Psicología": 2,
        "Economía": 2,
    }

    preguntas_por_curso_semillero = {
        "Razonamiento Verbal": 10,
        "Razonamiento Matemático": 10,
        "Aritmética": 5,
        "Álgebra": 5,
        "Literatura": 5,
        "Geometría": 0,
        "Trigonometría": 0,
        "Física": 0,
        "Química": 0,
        "Biología": 0,
        "Lenguaje": 0,
        "Historia": 0,
        "Geografía": 0,
        "Filosofía": 0,
        "Psicología": 0,
        "Economía": 0,
    }

    ruta_base = r'media'
    graficos = []
    fechas = set()

    for estudiante in Estudiante.objects.filter(reporte_actulizado = True):
        reportes = Reporte.objects.filter(KK_usuario=estudiante).order_by('-fecha_de_examen')

        for reporte in reportes:
            nivel = reporte.nivel
            if nivel == "90":
                cursos = cursos_pre
                preguntas_por_curso = preguntas_por_curso_pre
            else:
                cursos = cursos_semillero
                preguntas_por_curso = preguntas_por_curso_semillero
                

            datos = reporte.obtener_datos()
            carpeta_usuario = os.path.join(ruta_base, f"{estudiante.usuario}_2025", "imagenes_reporte")
            os.makedirs(carpeta_usuario, exist_ok=True)

            for curso in cursos:
                if curso not in datos:
                    continue  # Si no hay datos para ese curso, lo omitimos
                correctas, incorrectas = datos[curso]
                total_preguntas = preguntas_por_curso.get(curso, 10)
                en_blanco = total_preguntas - (correctas + incorrectas)

                fig, ax = plt.subplots(figsize=(8, 4))
                ax.bar(["Correctas", "Incorrectas", "En blanco"],
                       [correctas, incorrectas, en_blanco],
                       color=["green", "red", "gray"])
                ax.set_title(f"{curso} - {reporte.fecha_de_examen}")

                nombre_archivo = f"{curso}_{estudiante.usuario}_{reporte.fecha_de_examen}.png".replace(" ", "_")
                ruta_archivo = os.path.join(carpeta_usuario, nombre_archivo)

                plt.savefig(ruta_archivo, format='png')

                buffer = BytesIO()
                plt.savefig(buffer, format='png')
                buffer.seek(0)
                image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                buffer.close()
                plt.close(fig)

                graficos.append({
                    "fecha": reporte.fecha_de_examen,
                    "curso": curso,
                    "grafico": image_base64
                })

            fechas.add(reporte.fecha_de_examen)

    return render(request, 'ultimo_simulacro.html', {
        'graficos': graficos,
        'fechas': sorted(fechas),
        "base_template": base_template
    })


@login_required
@datos_actualizados_requerido('actualizar_datos')
def reportes_estudiante(request):
    usuario_actual = request.user

    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"

    reportes = Reporte.objects.filter(KK_usuario=usuario_actual).order_by('-fecha_de_examen')

    cursos = [
        "Razonamiento Verbal", "Razonamiento Matemático", "Aritmética", "Álgebra",
        "Geometría", "Trigonometría", "Física", "Química", "Biología", "Lenguaje",
        "Literatura", "Historia", "Geografía", "Filosofía", "Psicología", "Economía"
    ]

    fechas = list(reportes.values_list('fecha_de_examen', flat=True).distinct())
    graficos = []

    # Ruta base dentro de MEDIA_ROOT
    carpeta_usuario = os.path.join(settings.MEDIA_ROOT, f"{usuario_actual.usuario}_2025")

    for reporte in reportes:
        datos = reporte.obtener_datos()

        for curso in datos.keys():
            # Generar el nombre del archivo según la estructura "curso_usuario_fecha.png"
            nombre_archivo = f"{curso}_{usuario_actual.usuario}_{reporte.fecha_de_examen}.png".replace(" ", "_")
            ruta_relativa = os.path.join(f"{usuario_actual.usuario}_2025", nombre_archivo)
            ruta_absoluta = os.path.join(settings.MEDIA_ROOT, ruta_relativa)

            # Verificar si el archivo realmente existe antes de agregarlo
            if os.path.exists(ruta_absoluta):
                graficos.append({
                    "fecha": reporte.fecha_de_examen,
                    "curso": curso,
                    "ruta_imagen": f"{settings.MEDIA_URL}{ruta_relativa.replace('\\', '/')}"
                })
    
    return render(request, 'home.html', {
        'graficos': graficos,
        'cursos': cursos,
        'fechas': fechas,
        'base_template': base_template,
        'usuario_actual': usuario_actual,  # <--- esto
    })

@login_required
@datos_actualizados_requerido('actualizar_datos')
def reportes_observaciones(request):
    usuario_actual = request.user
    if usuario_actual.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular

    reportes = Reporte.objects.filter(KK_usuario=usuario_actual).order_by('-fecha_de_examen')

    observaciones = [
        {"fecha": reporte.fecha_de_examen.strftime('%Y-%m-%d'), "texto": reporte.Observacion}
        for reporte in reportes
    ]

    return render(request, 'observaciones.html', {'observaciones': observaciones, "base_template": base_template})

@login_required
@datos_actualizados_requerido('actualizar_datos')
def reportes_puesto_puntaje(request):
    usuario_actual = request.user
    if usuario_actual.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    elif usuario_actual.tipo_estudiante == "estudiante":
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular
    elif usuario_actual.tipo_estudiante == "profesor":
        base_template = "layouts/base_profesor.html"
    else:
        base_template = "layouts/base_tutor.html"

    reportes = Reporte.objects.filter(KK_usuario=usuario_actual).order_by('fecha_de_examen')

    fechas = [reporte.fecha_de_examen.strftime('%d-%m') for reporte in reportes]
    puestos = [reporte.puesto for reporte in reportes]
    puntajes = [reporte.obtener_total_puntaje() for reporte in reportes]

    def generar_grafico_puntaje(x, y, titulo, xlabel, ylabel, color, invertir_y=False):
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.plot(x, y, marker='o', linestyle='-', color=color)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)

        if invertir_y:
            ax.invert_yaxis()

        # Mostrar valores con dos decimales sobre los puntos
        for i, valor in enumerate(y):
            ax.annotate(f'{valor:.2f}', (x[i], y[i]),  # <- redondeo aquí
                        textcoords="offset points",
                        xytext=(0, 8),
                        ha='center',
                        fontsize=15,
                        color=color)

        # Rotar las fechas en el eje X
        plt.xticks(rotation=90)
        fig.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        imagen_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        buffer.close()
        plt.close(fig)

        return imagen_base64
    

    def generar_grafico_puesto(x, y, titulo, xlabel, ylabel, color, invertir_y=False):
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.plot(x, y, marker='o', linestyle='-', color=color)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)

        if invertir_y:
            ax.invert_yaxis()

        for i, valor in enumerate(y):
            ax.annotate(str(valor), (x[i], y[i]),
                        textcoords="offset points",
                        xytext=(0, 8),
                        ha='center',
                        fontsize=15,
                        color=color)

        # Rotar etiquetas del eje x y ajustar layout
        plt.xticks(rotation=90)
        fig.tight_layout()  # <- Esto asegura que no se corten las etiquetas

        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        imagen_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        buffer.close()
        plt.close(fig)

        return imagen_base64


    grafico_puesto = generar_grafico_puesto(fechas, puestos, "Evolución del Puesto", "Fecha", "Puesto", "blue", invertir_y=True)
    grafico_puntaje = generar_grafico_puntaje(fechas, puntajes, "Evolución del Puntaje", "Fecha", "Puntaje", "green")

    # Obtener nombre y último puntaje
    nombre_usuario = usuario_actual.nombre or usuario_actual.username
    ultimo_puntaje = puntajes[-1] if puntajes else None

    return render(request, 'ultimo_simulacro.html', {
        'grafico_puesto': grafico_puesto,
        'grafico_puntaje': grafico_puntaje,
        'base_template': base_template,
        'nombre_usuario': nombre_usuario,
        'ultimo_puntaje': ultimo_puntaje,
    })


#Administración de estudiantes

def registrar_asistencia(request):
    usuario_actual = request.user
    if usuario_actual.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"
    else:
        base_template = "layouts/base2.html"
    if request.method == 'POST':
        form = AsistenciaForm2(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_asistencias')  # o cualquier vista a la que quieras volver
    else:
        form = AsistenciaForm2()
    return render(request, 'registrar_asistencia.html', {'base_template': base_template,'form': form})

@login_required
@estudiante_tipo_requerido(['administrador'])
def ver_todas_asistencias(request):
    fecha_filtrada = request.GET.get('fecha')
    query = request.GET.get('q', '').strip()

    asistencias_qs = Asistencia.objects.all().select_related('KK_usuario').order_by('-Fecha', '-Hora')
    ultima_asistencia_presencial = asistencias_qs.filter(Modalidad='PRESENCIAL').first()
    ultima_fecha = None

    if ultima_asistencia_presencial:
        try:
            fecha = (ultima_asistencia_presencial.Fecha 
                    if not isinstance(ultima_asistencia_presencial.Fecha, str) 
                    else datetime.strptime(ultima_asistencia_presencial.Fecha, "%Y-%m-%d").date())

            hora = (
                    ultima_asistencia_presencial.Hora
                    if not isinstance(ultima_asistencia_presencial.Hora, str)
                    else datetime.strptime(ultima_asistencia_presencial.Hora, "%H:%M").time()
)

            ultima_fecha = f"{fecha.strftime('%d/%m/%Y')} a las {hora.strftime('%H:%M')}"
        except Exception as e:
            ultima_fecha = None  # En caso de error, lo ocultamos

    # Aplica filtros antes del slice
    if fecha_filtrada:
        try:
            fecha_obj = datetime.strptime(fecha_filtrada, "%Y-%m-%d").date()
            asistencias_qs = asistencias_qs.filter(Fecha=fecha_obj)
        except ValueError:
            pass

    if query:
        asistencias_qs = asistencias_qs.filter(
            Q(KK_usuario__nombre__icontains=query) |
            Q(KK_usuario__usuario__icontains=query)
        )

    # Aplica el slice después de los filtros
    asistencias = asistencias_qs[:200]

    # Ahora extrae las fechas distintas desde los registros ya limitados
    fechas = sorted(set(a.Fecha for a in asistencias))

    datos = []
    for fecha in fechas:
        marcas_fecha = [a for a in asistencias if a.Fecha == fecha]
        marcas_fecha.sort(key=lambda a: (a.KK_usuario.usuario, a.Hora))
        marcas_por_usuario = defaultdict(list)

        for asistencia in marcas_fecha:
            marcas_por_usuario[asistencia.KK_usuario.usuario].append(asistencia)

        for usuario, marcas in marcas_por_usuario.items():
            for i, asistencia in enumerate(marcas, start=1):
                datos.append({
                    "id": asistencia.ID_Reporte,
                    "usuario": asistencia.KK_usuario.usuario,
                    "nombre": asistencia.KK_usuario.nombre,
                    "numero_marca": i,
                    "fecha": fecha.strftime('%d/%m/%Y'),
                    "fecha_corta": dias_semana[fecha.weekday()] + '/' + fecha.strftime('%d/%m/%Y'),
                    "marca": "ENTRADA" if i in [1, 3] else "SALIDA",
                    "hora": asistencia.Hora,
                    "modalidad": asistencia.Modalidad,
                    "observacion": asistencia.Observacion
                })


    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string("fragmentos/tabla_asistencias.html", {"asistencias": datos})
        return JsonResponse({"html": html})

    return render(request, "listar_asistencias.html", {
        "asistencias": datos,
        "fecha_filtrada": fecha_filtrada,
        "query": query,
        "ultima_fecha": ultima_fecha,
        "base_template": "layouts/base.html"
    })

@login_required
@estudiante_tipo_requerido(['administrador'])
def eliminar_asistencia(request, pk):
    asistencia = get_object_or_404(Asistencia, pk=pk)
    asistencia.delete()
    messages.success(request, "La asistencia fue eliminada correctamente.")
    return redirect('listar_asistencias')  # Asegúrate de tener esta URL nombrada

@login_required
@estudiante_tipo_requerido(['administrador'])
def editar_asistencia(request, pk):
    usuario_actual = request.user

    # Definir la plantilla según el tipo de usuario
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    asistencia = get_object_or_404(Asistencia, pk=pk)

    if request.method == 'POST':
        nueva_fecha = request.POST.get('fecha')
        nueva_modalidad = request.POST.get('modalidad')
        nueva_observacion = request.POST.get('observacion')

        asistencia.Fecha = nueva_fecha
        asistencia.Modalidad = nueva_modalidad
        asistencia.Observacion = nueva_observacion
        asistencia.save()

        messages.success(request, 'Asistencia actualizada correctamente.')
        return redirect('listar_asistencias')

    return render(request, 'editar_asistencia.html', {'base_template': base_template,
        'asistencia': asistencia
    })

@login_required
@estudiante_tipo_requerido(['administrador'])
def acciones_asistencias(request):
    if request.method == 'POST':
        ids = request.POST.getlist('asistencias')
        accion = request.POST.get('accion')

        if not ids:
            messages.warning(request, "No seleccionaste asistencias.")
            return redirect('listar_asistencias')

        asistencias = Asistencia.objects.filter(pk__in=ids)

        if accion == 'eliminar':
            cantidad = asistencias.count()
            asistencias.delete()
            messages.success(request, f"{cantidad} asistencia(s) eliminada(s) correctamente.")
            return redirect('listar_asistencias')

        elif accion == 'descargar':
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(A4))
            width, height = landscape(A4)

            def dibujar_encabezado(y_pos):
                p.setFont("Helvetica-Bold", 14)
                p.drawString(30, height - 40, "Reporte de Asistencias - Academia Robert Hooke")

                headers = ["Usuario", "Nombre", "Fecha", "Hora", "Nro Marca", "Modalidad", "Observación"]
                x_positions = [30, 80, 310, 390, 470, 550, 650]
                p.setFont("Helvetica-Bold", 10)
                for i, header in enumerate(headers):
                    p.drawString(x_positions[i], y_pos, header)
                return y_pos - 25

            y = dibujar_encabezado(height - 80)
            p.setFont("Helvetica", 10)

            # Agrupamos asistencias por (usuario, fecha)
            from collections import defaultdict
            agrupadas = defaultdict(list)
            for a in asistencias.order_by('KK_usuario__usuario', 'Fecha', 'Hora'):
                clave = (a.KK_usuario.usuario, a.Fecha)
                agrupadas[clave].append(a)

            for grupo in agrupadas.values():
                for i, a in enumerate(grupo, start=1):
                    datos = [
                        a.KK_usuario.usuario,
                        a.KK_usuario.nombre,
                        a.Fecha.strftime('%d/%m/%Y'),
                        a.Hora,
                        "ENTRADA" if i in [1, 3] else "SALIDA",
                        a.Modalidad,
                        a.Observacion or "-"
                    ]
                    x_positions = [30, 80, 310, 390, 470, 550, 650]
                    for j, valor in enumerate(datos):
                        p.drawString(x_positions[j], y, str(valor))
                    y -= 20

                    if y < 40:
                        p.showPage()
                        y = dibujar_encabezado(height - 50)
                        p.setFont("Helvetica", 10)

            p.save()
            buffer.seek(0)

            return FileResponse(buffer, as_attachment=True, filename="asistencias_seleccionadas.pdf")


    return redirect('listar_asistencias')


@require_POST
@login_required
@estudiante_tipo_requerido(['administrador'])
def eliminar_estudiantes_masivo(request):
    ids = request.POST.getlist('estudiantes')
    ids = [id for id in ids if id.isdigit()]
    
    if ids:
        estudiantes = Estudiante.objects.filter(ID_Estudiante__in=ids)
        
        for estudiante in estudiantes:
            # Ruta a la carpeta del estudiante
            carpeta_estudiante = os.path.join(settings.MEDIA_ROOT, f"{estudiante.usuario}_2025")
            
            reporte_estudiante = os.path.join(settings.MEDIA_ROOT, f"reportes/simulacros/{estudiante.usuario}_reporte_simulacro.png")

            asistencia_estudiante = os.path.join(settings.MEDIA_ROOT, f"reportes/asistencias/{estudiante.usuario}.png")

            # Eliminar carpeta si existe
            if os.path.exists(carpeta_estudiante) and os.path.isdir(carpeta_estudiante):
                shutil.rmtree(carpeta_estudiante)

            if os.path.exists(reporte_estudiante) and os.path.isfile(reporte_estudiante):
                os.remove(reporte_estudiante)
            
            if os.path.exists(asistencia_estudiante) and os.path.isfile(asistencia_estudiante):
                os.remove(asistencia_estudiante)
        # Eliminar los estudiantes después de borrar las carpetas
        estudiantes.delete()
        
        messages.success(request, "Estudiantes y carpetas eliminados correctamente.")
    else:
        messages.warning(request, "No seleccionaste ningún estudiante válido.")
    
    return redirect('lista_estudiantes')

@login_required
@estudiante_tipo_requerido(['administrador', 'tutor'])
def agregar_observacion(request, estudiante_id):
    usuario_actual = request.user


    if usuario_actual.tipo_estudiante == 'administrador':
        url_volver = reverse('lista_estudiantes')
    else:
        url_volver = reverse('tutorados')

    if usuario_actual.tipo_estudiante == "administrador":
        base_template = "layouts/base.html"
    elif usuario_actual.tipo_estudiante == "tutor":
        base_template = "layouts/base_tutor.html"
    else:
        base_template = "layouts/base2.html"
    # Definir la plantilla según el tipo de usuario
    estudiante = get_object_or_404(Estudiante, ID_Estudiante=estudiante_id)
    reportes = Reporte.objects.filter(KK_usuario=estudiante)
    
    fecha_str = request.GET.get('fecha')
    fecha_obj = None

    if fecha_str:
        try:
            # Forzar que la fecha esté en formato YYYY-MM-DD
            fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponseBadRequest("Formato de fecha inválido")


    if request.method == 'POST':
        observacion = request.POST.get('observacion')
        if not fecha_obj:
            return HttpResponseBadRequest("Falta la fecha del examen")

        reporte = reportes.filter(fecha_de_examen=fecha_obj).first()
        if reporte:
            reporte.Observacion = observacion
            reporte.save()
            if usuario_actual.tipo_estudiante == 'administrador':
                return redirect('lista_estudiantes')  # o a donde quieras redirigir
            else:
                return redirect('tutorados')
        else:
            return HttpResponseBadRequest("No se encontró el reporte para esa fecha")


    return render(request, 'agregar_observacion.html', {
        'url_volver': url_volver,
        'estudiante': estudiante,
        'reportes': reportes,
        'base_template': base_template,
        'fecha_seleccionada': fecha_obj,
        'reporte_seleccionado': reportes.filter(fecha_de_examen=fecha_obj).first() if fecha_obj else None
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def lista_estudiantes(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    
    query = request.GET.get('q', '').strip()

    if query:
        estudiantes = Estudiante.objects.filter(
            Q(nombre__icontains=query) | Q(usuario__icontains=query),
            tipo_estudiante="estudiante"
        )
    else:
        estudiantes = Estudiante.objects.filter(tipo_estudiante="estudiante")

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string("fragmentos/tabla_estudiantes.html", {"estudiantes": estudiantes})
        return JsonResponse({"html": html})

    return render(request, 'lista_estudiantes.html', {
        'estudiantes': estudiantes,
        'query': query,
        "base_template": base_template
    })



@login_required
@estudiante_tipo_requerido(['administrador'])
def lista_estudiantes_respaldo(request):
    usuario_actual = request.user

    # Definir la plantilla según el tipo de usuario
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    query = request.GET.get('q', '')
    if query:
        estudiantes = Estudiante.objects.filter(nombre__icontains=query)
    else:
        estudiantes = Estudiante.objects.filter(tipo_estudiante = "estudiante")
    return render(request, 'lista_estudiantes.html', {
        'estudiantes': estudiantes,
        'query': query, 
        "base_template": base_template
    })

@login_required
@estudiante_tipo_requerido(['administrador'])
def editar_estudiante(request, pk):
    usuario_actual = request.user

    # Definir la plantilla según el tipo de usuario
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"
    estudiante = get_object_or_404(Estudiante, pk=pk)
    if request.method == 'POST':
        form = EstudianteForm2(request.POST, instance=estudiante)
        if form.is_valid():
            form.save()
            return redirect('lista_estudiantes')
    else:
        form = EstudianteForm2(instance=estudiante)
    return render(request, 'editar_estudiante.html', {'form': form, "base_template": base_template})


@login_required
@estudiante_tipo_requerido(['administrador'])
def eliminar_estudiante(request, pk):
    estudiante = get_object_or_404(Estudiante, pk=pk)

    # Elimina sin confirmación HTML
    ruta_base = r'media'
    carpeta_estudiante = os.path.join(ruta_base, f"{estudiante.usuario}_2025")

    if os.path.exists(carpeta_estudiante) and os.path.isdir(carpeta_estudiante):
        shutil.rmtree(carpeta_estudiante)

    estudiante.delete()
    return redirect('lista_estudiantes')


@login_required
@estudiante_tipo_requerido(['administrador'])
def eliminar_estudiante_respaldo(request, pk):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"

    estudiante = get_object_or_404(Estudiante, pk=pk)

    if request.method == 'POST':
        # Ruta base donde están las carpetas de los estudiantes
        ruta_base = r'media'

        # Nombre de la carpeta del estudiante: "usuario2025"
        carpeta_estudiante = os.path.join(ruta_base, f"{estudiante.usuario}_2025")

        # Verifica si la carpeta existe y elimínala
        if os.path.exists(carpeta_estudiante) and os.path.isdir(carpeta_estudiante):
            shutil.rmtree(carpeta_estudiante)  # Elimina la carpeta completa y su contenido

        # Luego elimina el estudiante de la base de datos
        estudiante.delete()
        return redirect('lista_estudiantes')

    return render(request, 'eliminar_estudiante.html', {'estudiante': estudiante, "base_template": base_template})


dias_semana = {
    0: "LUNES",
    1: "MARTES",
    2: "MIÉRCOLES",
    3: "JUEVES",
    4: "VIERNES",
    5: "SÁBADO",
    6: "DOMINGO",
}

@login_required
@datos_actualizados_requerido('actualizar_datos')
def ver_asistencias(request):
    estudiante = request.user
    base_template = "layouts/base.html" if estudiante.tipo_estudiante == "administrador" else "layouts/base2.html"
    fecha_filtrada = request.GET.get('fecha')
    # Filtrar asistencias del estudiante y ordenarlas
    asistencias = Asistencia.objects.filter(KK_usuario=estudiante).order_by('Fecha', 'Hora')

    if fecha_filtrada:
        try:
            fecha_obj = datetime.strptime(fecha_filtrada, "%Y-%m-%d").date()
            asistencias = asistencias.filter(Fecha=fecha_obj)
        except ValueError:
            pass  # Fecha inválida, ignora el filtro

    agrupadas_por_fecha = defaultdict(list)
    for asistencia in asistencias:
        agrupadas_por_fecha[asistencia.Fecha].append(asistencia)

    datos = []
    for fecha in asistencias.values_list('Fecha', flat=True).distinct().order_by('Fecha'):
        marcas = asistencias.filter(Fecha=fecha).order_by('Hora')
        for i, asistencia in enumerate(marcas, start=1):
            dia_semana = dias_semana[fecha.weekday()]
            fecha_corta = f"{dia_semana} {fecha.strftime('%d/%m')}"
            datos.append({
                "usuario": estudiante.usuario,
                "nombre": estudiante.nombre,
                "numero_marca": i,
                "fecha": fecha.strftime('%d/%m/%Y'),
                "fecha_corta": fecha_corta,
                "hora": asistencia.Hora,
                "observacion": asistencia.Observacion,
                "modalidad" : asistencia.Modalidad,
            })

    return render(request, "ver_asistencias.html", {"asistencias": datos, "base_template": base_template, "fecha_filtrada": fecha_filtrada})


@login_required
@estudiante_tipo_requerido(['administrador'])
def cargar_asistencias(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"

    if request.method == 'POST':
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            try:
                df = pd.read_excel(archivo_excel)
                columnas_esperadas = ['usuario', 'marca', 'modalidad']
                if not all(col in df.columns for col in columnas_esperadas):
                    messages.error(request, "El archivo Excel no tiene las columnas esperadas: 'usuario', 'marca' y 'modalidad'")
                    return redirect('subir_asistencia')

                marcas_por_estudiante_fecha = defaultdict(list)

                for _, fila in df.iterrows():
                    try:
                        estudiante = Estudiante.objects.get(usuario=fila['usuario'])
                        marca_datetime = pd.to_datetime(fila['marca'])
                        fecha = marca_datetime.date()
                        modalidad = fila.get('modalidad')
                        marcas_por_estudiante_fecha[(estudiante, fecha)].append((marca_datetime, modalidad))
                    except Estudiante.DoesNotExist:
                        messages.warning(request, f"El usuario '{fila['usuario']}' no existe en la base de datos.")
                    except Exception as e:
                        messages.warning(request, f"Error procesando fila: {str(e)}")

                config = VariableControl.objects.get(ID_Variable=1)

                for (estudiante, fecha), marcas in marcas_por_estudiante_fecha.items():
                    franjas = {
                        'mañana': (config.EntradaManana, config.SalidaManana),
                        'tarde': (config.EntradaTarde, config.SalidaTarde),
                        'amanecida': (config.EntradaAmanecida, config.SalidaAmanecida),
                    }

                    for franja_nombre, (hora_inicio, hora_fin) in franjas.items():
                        marcas_franja = [m for m in marcas if hora_inicio <= m[0].time() <= hora_fin]

                        if not marcas_franja:
                            continue

                        min_marca = min(marcas_franja, key=lambda x: x[0])
                        max_marca = max(marcas_franja, key=lambda x: x[0])

                        # Traer registros existentes de esta franja
                        registros_existentes = Asistencia.objects.filter(
                            KK_usuario=estudiante,
                            Fecha=fecha,
                            Hora__gte=hora_inicio.strftime("%H:%M"),
                            Hora__lte=hora_fin.strftime("%H:%M")
                        ).order_by("Hora")

                        # Si no hay registros aún
                        if registros_existentes.count() < 2:
                            if registros_existentes.filter(Hora=min_marca[0].strftime("%H:%M")).exists() == False:
                                Asistencia.objects.create(
                                    KK_usuario=estudiante,
                                    Fecha=fecha,
                                    Hora=min_marca[0].strftime("%H:%M"),
                                    Modalidad=min_marca[1]
                                )
                            if min_marca != max_marca and registros_existentes.filter(Hora=max_marca[0].strftime("%H:%M")).exists() == False:
                                Asistencia.objects.create(
                                    KK_usuario=estudiante,
                                    Fecha=fecha,
                                    Hora=max_marca[0].strftime("%H:%M"),
                                    Modalidad=max_marca[1]
                                )
                        else:
                            # Ya hay dos registros. Verificar si reemplazar alguno.
                            hora_min_existente = min(registros_existentes, key=lambda r: r.Hora)
                            hora_max_existente = max(registros_existentes, key=lambda r: r.Hora)

                            nueva_hora_min = min_marca[0].strftime("%H:%M")
                            nueva_hora_max = max_marca[0].strftime("%H:%M")

                            # Reemplazar min si la nueva hora es menor
                            if nueva_hora_min < hora_min_existente.Hora:
                                hora_min_existente.delete()
                                Asistencia.objects.create(
                                    KK_usuario=estudiante,
                                    Fecha=fecha,
                                    Hora=nueva_hora_min,
                                    Modalidad=min_marca[1]
                                )

                            # Reemplazar max si la nueva hora es mayor
                            if nueva_hora_max > hora_max_existente.Hora:
                                hora_max_existente.delete()
                                Asistencia.objects.create(
                                    KK_usuario=estudiante,
                                    Fecha=fecha,
                                    Hora=nueva_hora_max,
                                    Modalidad=max_marca[1]
                                )


                messages.success(request, "Datos cargados exitosamente.")
                return redirect('subir_asistencia')

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('subir_asistencia')
    else:
        form = CargarExcelForm()

    return render(request, "subir_asistencia.html", {"base_template": base_template, "form": form})

@login_required
@estudiante_tipo_requerido(['administrador'])
def cargar_asistencias_respalds(request):
    estudiante = request.user
    base_template = "layouts/base.html" if estudiante.tipo_estudiante == "administrador" else "layouts/base2.html"

    if request.method == 'POST':
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            try:
                df = pd.read_excel(archivo_excel)
                columnas_esperadas = ['usuario', 'marca', 'modalidad']
                if not all(col in df.columns for col in columnas_esperadas):
                    messages.error(request, "El archivo Excel no tiene las columnas esperadas: 'usuario', 'marca' y 'modalidad'")
                    return redirect('subir_asistencia')

                # Agrupar marcas por usuario y fecha
                marcas_por_estudiante_fecha = defaultdict(list)

                for _, fila in df.iterrows():
                    try:
                        estudiante = Estudiante.objects.get(usuario=fila['usuario'])
                        marca_datetime = pd.to_datetime(fila['marca'])
                        fecha = marca_datetime.date()
                        modalidad = fila.get('modalidad')
                        marcas_por_estudiante_fecha[(estudiante, fecha)].append((marca_datetime, modalidad))
                    except Estudiante.DoesNotExist:
                        messages.warning(request, f"El usuario '{fila['usuario']}' no existe en la base de datos.")
                    except Exception as e:
                        messages.warning(request, f"Error procesando fila: {str(e)}")

                config = VariableControl.objects.get(ID_Variable=1)

                HoraEntradaManana = config.EntradaManana
                HoraSalidaManana = config.SalidaManana
                HoraEntradaTarde = config.EntradaTarde
                HoraSalidaTarde = config.SalidaTarde
                HoraEntradaAmanecida = config.EntradaAmanecida
                HoraSalidaAmanecida = config.SalidaAmanecida

                for (estudiante, fecha), marcas in marcas_por_estudiante_fecha.items():
                    # Separar en horarios de manana, tarde y amanecida
                    manana = [m for m in marcas if HoraEntradaManana <= m[0].time() <= HoraSalidaManana]
                    tarde = [m for m in marcas if HoraEntradaTarde <= m[0].time() <= HoraSalidaTarde]
                    amanecida = [m for m in marcas if HoraEntradaAmanecida <= m[0].time() <= HoraSalidaAmanecida]

                    registros = []

                    if manana:
                        entrada_manana = min(manana, key=lambda x: x[0])
                        salida_manana = max(manana, key=lambda x: x[0])
                        if entrada_manana[0] != salida_manana[0]:
                            registros.append(entrada_manana)
                            registros.append(salida_manana)
                        else:
                            registros.append(entrada_manana)

                    if tarde:
                        entrada_tarde = min(tarde, key=lambda x: x[0])
                        salida_tarde = max(tarde, key=lambda x: x[0])
                        if entrada_tarde[0] != salida_tarde[0]:
                            registros.append(entrada_tarde)
                            registros.append(salida_tarde)
                        else:
                            registros.append(entrada_tarde)
                    
                    if amanecida:
                        entrada_amanecida = min(amanecida, key=lambda x: x[0])
                        salida_amanecida = max(amanecida, key=lambda x: x[0])
                        if entrada_amanecida[0] != salida_amanecida[0]:
                            registros.append(entrada_amanecida)
                            registros.append(salida_amanecida)
                        else:
                            registros.append(entrada_amanecida)

                    for marca_datetime, modalidad in registros:
                        hora = marca_datetime.strftime("%H:%M")
                        Asistencia.objects.create(
                            KK_usuario=estudiante,
                            Fecha=fecha,
                            Hora=hora,
                            Modalidad=modalidad
                        )

                messages.success(request, "Datos cargados exitosamente.")
                return redirect('subir_asistencia')

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('subir_asistencia')
    else:
        form = CargarExcelForm()

    return render(request, "subir_asistencia.html", {"base_template": base_template, "form": form})


@login_required
@estudiante_tipo_requerido(['administrador'])
def cargar_asistencias_respaldo_TMA(request):
    estudiante = request.user
    base_template = "layouts/base.html" if estudiante.tipo_estudiante == "administrador" else "layouts/base2.html"

    if request.method == 'POST':
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            try:
                df = pd.read_excel(archivo_excel)
                columnas_esperadas = ['usuario', 'marca', 'modalidad']
                if not all(col in df.columns for col in columnas_esperadas):
                    messages.error(request, "El archivo Excel no tiene las columnas esperadas: 'usuario', 'marca' y 'modalidad'")
                    return redirect('subir_asistencia')

                # Agrupar marcas por usuario y fecha
                marcas_por_estudiante_fecha = defaultdict(list)

                for _, fila in df.iterrows():
                    try:
                        estudiante = Estudiante.objects.get(usuario=fila['usuario'])
                        marca_datetime = pd.to_datetime(fila['marca'])
                        fecha = marca_datetime.date()
                        modalidad = fila.get('modalidad')
                        marcas_por_estudiante_fecha[(estudiante, fecha)].append((marca_datetime, modalidad))
                    except Estudiante.DoesNotExist:
                        messages.warning(request, f"El usuario '{fila['usuario']}' no existe en la base de datos.")
                    except Exception as e:
                        messages.warning(request, f"Error procesando fila: {str(e)}")

                # Obtener los horarios desde la base de datos
                config = VariableControl.objects.get(ID_Variable=1)

                HoraEntradaManana = config.EntradaManana
                HoraSalidaManana = config.SalidaManana
                HoraEntradaTarde = config.EntradaTarde
                HoraSalidaTarde = config.SalidaTarde
                HoraEntradaAmanecida = config.EntradaAmanecida
                HoraSalidaAmanecida = config.SalidaAmanecida

                for (estudiante, fecha), marcas in marcas_por_estudiante_fecha.items():
                    registros = []

                    # Separar marcas por turno
                    manana = [m for m in marcas if HoraEntradaManana <= m[0].time() <= HoraSalidaManana]
                    tarde = [m for m in marcas if HoraEntradaTarde <= m[0].time() <= HoraSalidaTarde]
                    amanecida = [m for m in marcas if HoraEntradaAmanecida <= m[0].time() <= HoraSalidaAmanecida]

                    # Mañana
                    if manana:
                        entrada = min(manana, key=lambda x: x[0])
                        salida = max(manana, key=lambda x: x[0])
                        if entrada[0] != salida[0]:
                            registros.append((entrada[0]))
                            registros.append((salida[0]))
                        else:
                            registros.append((entrada[0]))

                    # Tarde
                    if tarde:
                        entrada = min(tarde, key=lambda x: x[0])
                        salida = max(tarde, key=lambda x: x[0])
                        if entrada[0] != salida[0]:
                            registros.append((entrada[0]))
                            registros.append((salida[0]))
                        else:
                            registros.append((entrada[0]))

                    # Amanecida
                    if amanecida:
                        entrada = min(amanecida, key=lambda x: x[0])
                        salida = max(amanecida, key=lambda x: x[0])
                        if entrada[0] != salida[0]:
                            registros.append((entrada[0]))
                            registros.append((salida[0]))
                        else:
                            registros.append((entrada[0]))

                    # Guardar registros
                    for marca_datetime, modalidad, observacion in registros:
                        Asistencia.objects.create(
                            KK_usuario=estudiante,
                            Fecha=fecha,
                            Hora=marca_datetime.strftime("%H:%M"),
                            Modalidad=modalidad,
                            Observacion=observacion
                        )
                messages.success(request, "Datos cargados exitosamente.")
                return redirect('subir_asistencia')

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('subir_asistencia')
    else:
        form = CargarExcelForm()

    return render(request, "subir_asistencia.html", {"base_template": base_template, "form": form})





@login_required
@estudiante_tipo_requerido(['administrador'])
def subir_reporte(request):
    estudiante = request.user
    base_template = "layouts/base.html" if estudiante.tipo_estudiante == "administrador" else "layouts/base2.html"

    if request.method == 'POST':
        form = CargarExcelFormReporte(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            nivel = int(form.cleaned_data['nivel'])  # 30 o 90

            try:
                df = pd.read_excel(archivo_excel)

                columnas_semi = ['Alumno', 'Fecha Simulacro', 'Puesto', 'Rv_1', 'Rv_2', 'Rm_1', 'Rm_2', 'Ar_1', 'Ar_2', 'Al_1', 'Al_2', 'Lit_1', 'Lit_2'
                ]
                columnas_pre = [
                    'Alumno', 'Fecha Simulacro', 'Puesto', 'Rv_1', 'Rv_2', 'Rm_1', 'Rm_2', 'Ar_1', 'Ar_2', 'Al_1', 'Al_2','Ge_1', 'Ge_2', 'Tr_1', 'Tr_2', 'Fi_1', 'Fi_2', 'Qu_1', 'Qu_2', 'Bi_1', 'Bi_2','Le_1', 'Le_2', 'Lit_1', 'Lit_2', 'Hi_1', 'Hi_2', 'Gf_1', 'Gf_2', 'Fil_1', 'Fil_2','Psi_1', 'Psi_2', 'Ec_1', 'Ec_2'
                ]

                # Validar columnas según nivel
                if nivel == 30 and not all(col in df.columns for col in columnas_semi):
                    messages.error(request, "El archivo no tiene las columnas esperadas para nivel Semillero, revise los campos por favor: 'Alumno', 'Fecha Simulacro', 'Puesto', 'Rv_1', 'Rv_2', 'Rm_1', 'Rm_2', 'Ar_1', 'Ar_2', 'Al_1', 'Al_2', 'Lit_1', 'Lit_2', 'Observacion'")
                    return redirect('subir_reporte')
                if nivel == 90 and not all(col in df.columns for col in columnas_pre):
                    messages.error(request, "El archivo no tiene las columnas esperadas para nivel Pre: 'Alumno', 'Fecha Simulacro', 'Puesto', 'Rv_1', 'Rv_2', 'Rm_1', 'Rm_2', 'Ar_1', 'Ar_2', 'Al_1', 'Al_2','Ge_1', 'Ge_2', 'Tr_1', 'Tr_2', 'Fi_1', 'Fi_2', 'Qu_1', 'Qu_2', 'Bi_1', 'Bi_2','Le_1', 'Le_2', 'Lit_1', 'Lit_2', 'Hi_1', 'Hi_2', 'Gf_1', 'Gf_2', 'Fil_1', 'Fil_2','Psi_1', 'Psi_2', 'Ec_1', 'Ec_2', 'Observacion'")
                    return redirect('subir_reporte')
                actualizados = 0
                nuevos = 0
                for _, fila in df.iterrows():
                    try:
                        estudiante = Estudiante.objects.get(usuario=fila['Alumno'])
                        estudiante.reporte_actualizado = True
                        estudiante.save()

                        fecha_examen = fila.get('Fecha Simulacro')

                        # Intentar obtener un registro existente
                        reporte_existente = Reporte.objects.filter(KK_usuario=estudiante, fecha_de_examen=fecha_examen).first()

                        # Campos comunes
                        datos = {
                            "fecha_de_examen": fecha_examen,
                            "puesto": fila.get('Puesto'),
                            "nivel": nivel,
                            "Rv_1": fila.get("Rv_1", 0),
                            "Rv_2": fila.get("Rv_2", 0),
                            "Rm_1": fila.get("Rm_1", 0),
                            "Rm_2": fila.get("Rm_2", 0),
                            "Ar_1": fila.get("Ar_1", 0),
                            "Ar_2": fila.get("Ar_2", 0),
                            "Al_1": fila.get("Al_1", 0),
                            "Al_2": fila.get("Al_2", 0),
                            "Lit_1": fila.get("Lit_1", 0),
                            "Lit_2": fila.get("Lit_2", 0),
                            "reporte_actualizado": True,
                        }

                        if nivel == 90:
                            datos.update({
                                "Ge_1": fila.get("Ge_1", 0), "Ge_2": fila.get("Ge_2", 0),
                                "Tr_1": fila.get("Tr_1", 0), "Tr_2": fila.get("Tr_2", 0),
                                "Fi_1": fila.get("Fi_1", 0), "Fi_2": fila.get("Fi_2", 0),
                                "Qu_1": fila.get("Qu_1", 0), "Qu_2": fila.get("Qu_2", 0),
                                "Bi_1": fila.get("Bi_1", 0), "Bi_2": fila.get("Bi_2", 0),
                                "Le_1": fila.get("Le_1", 0), "Le_2": fila.get("Le_2", 0),
                                "Hi_1": fila.get("Hi_1", 0), "Hi_2": fila.get("Hi_2", 0),
                                "Gf_1": fila.get("Gf_1", 0), "Gf_2": fila.get("Gf_2", 0),
                                "Fil_1": fila.get("Fil_1", 0), "Fil_2": fila.get("Fil_2", 0),
                                "Psi_1": fila.get("Psi_1", 0), "Psi_2": fila.get("Psi_2", 0),
                                "Ec_1": fila.get("Ec_1", 0), "Ec_2": fila.get("Ec_2", 0),
                            })
                        else:
                            for campo in ['Ge', 'Tr', 'Fi', 'Qu', 'Bi', 'Le', 'Hi', 'Gf', 'Fil', 'Psi', 'Ec']:
                                datos[f"{campo}_1"] = 0
                                datos[f"{campo}_2"] = 0

                        if reporte_existente:
                            # Actualizar campos del reporte existente
                            for key, value in datos.items():
                                setattr(reporte_existente, key, value)
                            reporte_existente.save()
                            actualizados += 1
                        else:
                            # Crear nuevo reporte si no existe
                            Reporte.objects.create(KK_usuario=estudiante, **datos)
                            nuevos += 1

                    except Estudiante.DoesNotExist:
                        messages.error(request, f"El estudiante '{fila['Alumno']}' no existe.")
                        continue


                messages.success(request, f"Datos cargados exitosamente. {nuevos} nuevos y {actualizados} actualizados.")
                return redirect('subir_reporte')

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")

        else:
            messages.error(request, "Formulario inválido. Por favor, verifica los datos.")

    else:
        form = CargarExcelFormReporte()

    return render(request, "subir_reporte.html", {
        "base_template": base_template,
        "form": form
    })



@login_required
@estudiante_tipo_requerido(['administrador'])
def cargar_excel(request):
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular

    RUTA_BASE_ESTUDIANTES = r'media'

    if request.method == 'POST':
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            
            try:
                # Leer el archivo Excel
                df = pd.read_excel(archivo_excel)
                
                # Verificar que el archivo tenga las columnas correctas
                if not all(col in df.columns for col in ['nombre', 'usuario', 'contraseña']):
                    messages.error(request, 'El archivo Excel debe tener las columnas: nombre, usuario, contraseña.')
                    return redirect('cargar_excel')
                
                # Procesar cada fila del archivo
                for index, row in df.iterrows():
                    nombre = row['nombre']
                    usuario = row['usuario']
                    contraseña = row['contraseña']
                    
                    # Verificar si el usuario ya existe
                    if Estudiante.objects.filter(usuario=usuario).exists():
                        messages.warning(request, f'El usuario {usuario} ya existe y no fue registrado.')
                    else:
                        # Crear un nuevo estudiante
                        Estudiante.objects.create(
                            nombre=nombre,
                            usuario=usuario,
                            contraseña=contraseña
                        )
                        # Crear la carpeta del estudiante
                        carpeta_estudiante = os.path.join(RUTA_BASE_ESTUDIANTES, f"{usuario}_2025")

                        subcarpeta_imagenes = os.path.join(carpeta_estudiante, "imagenes_reporte")
                        if not os.path.exists(carpeta_estudiante):
                            os.makedirs(subcarpeta_imagenes, exist_ok=True)
        
                messages.success(request, f'Estudiante {usuario} registrado correctamente y carpeta creada.')
                
                messages.success(request, 'Archivo procesado correctamente.')
                return redirect('cargar_excel')
            
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
                return redirect('cargar_excel')
    else:
        form = CargarExcelForm()
    
    return render(request, "cargar_excel.html", {"base_template": base_template, "form": form})

@login_required
def crearAlumno(request):
    return render(request, 'crear_alumno.html')


def prueba(request):
    usuario_actual = request.user
    # Definir la plantilla según el tipo de usuario
    base_template = "layouts/base.html"
    return render(request, 'prueba.html', {
        'base_template': base_template,
    })


def login_view(request):
    user = request.user
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            usuario = form.cleaned_data['usuario']
            contraseña = form.cleaned_data['contraseña']
            
            # Autenticar usando el backend personalizado
            backend = EstudianteBackend()
            estudiante = backend.authenticate(request, username=usuario, password=contraseña)
            
            if estudiante is not None:
                login(request, estudiante, backend='tu_app.backends.EstudianteBackend')
                messages.success(request, 'Inicio de sesión exitoso.')

                # Redirigir según tipo de estudiante
                if estudiante.tipo_estudiante == "estudiante":
                    return redirect('home')
                elif estudiante.tipo_estudiante == "administrador":
                    return redirect('lista_estudiantes')
                else:
                    return redirect('login')  # Por si acaso

            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
    else:
        form = LoginForm()
    
    return render(request, 'registration/login.html', {'form': form})



def logout_view_respaldo(request):

    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular
    logout(request)

    return render(request, "home.html", {"base_template": base_template})


def login_view(request):
    user = request.user
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            usuario = form.cleaned_data['usuario']
            contraseña = form.cleaned_data['contraseña']
            
            # Autenticar usando el backend personalizado
            backend = EstudianteBackend()
            estudiante = backend.authenticate(request, username=usuario, password=contraseña)
            
            if estudiante is not None:
                login(request, estudiante, backend='tu_app.backends.EstudianteBackend')
                messages.success(request, 'Inicio de sesión exitoso.')
                return redirect('estudiantes')
            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
    else:
        form = LoginForm()
    
    return render(request, 'registration/login.html', {'form': form}, {'user': user})


@login_required
def actualizar_datos(request):
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":
        base_template = "layouts/base.html"
    elif estudiante.tipo_estudiante == "tutor":
        base_template = "layouts/base_tutor.html"
    elif estudiante.tipo_estudiante == "profesor":
        base_template = "layouts/base_profesor.html"
    else:
        base_template = "layouts/base2.html"
    
    estudiante = get_object_or_404(Estudiante, pk=estudiante.ID_Estudiante)
    if request.method == 'POST':
        if estudiante.tipo_estudiante == 'estudiante':
            form = ActualizarDatosForm(request.POST, instance=estudiante)
        else:
            form = ActualizarDatosFormProf(request.POST,instance=estudiante)
        if form.is_valid():
            estudiante.actualizado = "actualizado"
            form.save()
            return redirect(reverse('actualizar_datos') + '?exito=1')
    else:
        if estudiante.tipo_estudiante == 'estudiante':
            form = ActualizarDatosForm(instance=estudiante)
        else:
            form = ActualizarDatosFormProf(instance=estudiante)
    
    return render(request, 'actualizar_datos.html', {
        'base_template': base_template,
        'form': form,
        'exito': request.GET.get('exito') == '1'
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def registrar_estudiante(request):
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular

    if request.method == 'POST':
        form = EstudianteForm(request.POST)
        if form.is_valid():
            usuario = form.cleaned_data['usuario']
            if Estudiante.objects.filter(usuario=usuario).exists():
                messages.error(request, 'El usuario ya existe.')
            else:
                nuevo_estudiante = form.save()

                # Definir la ruta donde se crearán las carpetas
                ruta_base = r'media'
                
                # Crear la carpeta con el formato "usuario2025"
                carpeta_estudiante = os.path.join(ruta_base, f"{usuario}_2025")

                subcarpeta_imagenes = os.path.join(carpeta_estudiante, "imagenes_reporte")

                os.makedirs(subcarpeta_imagenes, exist_ok=True)

                messages.success(request, 'Estudiante creado correctamente y carpeta asignada.')
                return redirect('registrar_estudiante')
        else:
            messages.error(request, 'Error al crear el estudiante. Revise los datos.')
    else:
        form = EstudianteForm()

    return render(request, "registrar_estudiante.html", {"base_template": base_template, "form": form})

@login_required
def base(request):
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular

    return render(request, "registration/login.html", {"base_template": base_template})


@login_required
def home(request):
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    elif estudiante.tipo_estudiante == "estudiante":
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular
    elif estudiante.tipo_estudiante == "profesor":
        base_template = "layouts/base_profesor.html"
    else:
        base_template = "layouts/base_tutor.html"

    return render(request, "home.html", {"base_template": base_template})

def inicio(request):
    return render(request, 'registration/login.html');


def salir(request):
    logout(request)
    return redirect('home')


def user_login(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")
    return render(request, 'login.html')

@login_required
def upload_excel(request):
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        try:
            df = pd.read_excel(excel_file)
            for index, row in df.iterrows():
                username = row['Usuario']
                password = row['Contraseña']
            messages.success(request, "Archivo subido y usuarios creados exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
    return render(request, "upload_excel.html", {"base_template": base_template})





##############################################3

def crear_grafico_estudiante_curso_respaldo(estudiante, nombre_curso, datos):
    df = pd.DataFrame(datos)
    if df.empty:
        return

    df['Fecha Simulacro'] = pd.to_datetime(df['Fecha Simulacro'])
    df = df.sort_values('Fecha Simulacro')
    df['blanco'] = 10 - df['correctas'] - df['incorrectas']

    fechas = df['Fecha Simulacro'].unique()
    cols = 3
    rows = math.ceil(len(fechas) / cols)

    fig, axes = plt.subplots(rows, cols, figsize=(20, 5 * rows))
    axes = axes.flatten()

    for i, fecha in enumerate(fechas):
        datos_fecha = df[df['Fecha Simulacro'] == fecha]
        ax = axes[i]

        index = range(len(datos_fecha))
        bar_width = 0.2

        bars1 = ax.bar(index, datos_fecha['correctas'] * 0.85, bar_width, color='green')
        bars2 = ax.bar([i + bar_width for i in index], datos_fecha['incorrectas'] * 0.85, bar_width, color='red')
        bars3 = ax.bar([i + 2 * bar_width for i in index], datos_fecha['blanco'] * 0.85, bar_width, color='gray')

        for spine in ax.spines.values():
            spine.set_visible(False)

        ax.set_xticks([])
        ax.set_xlabel(fecha.strftime('%d/%m/%Y'), fontsize=30)
        ax.set_yticks([])

        for bars, original_values in zip([bars1, bars2, bars3], [datos_fecha['correctas'], datos_fecha['incorrectas'], datos_fecha['blanco']]):
            for bar, original_value in zip(bars, original_values):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, height + 0.03, f"{int(original_value)}", ha='center', fontsize=40, fontweight='bold')

    for j in range(i + 1, len(axes)):
        fig.delaxes(axes[j])

    plt.subplots_adjust(bottom=0.15)

    carpeta = f'media/{estudiante.usuario}_2025/imagenes_reporte/'
    os.makedirs(carpeta, exist_ok=True)
    img_path = os.path.join(carpeta, f"{nombre_curso.replace(' ', '_')}.png")
    plt.savefig(img_path, dpi=300)
    plt.close()



@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_imagenes_reportes_respaldo(request):
    usuario_actual = request.user
    base_template = "layouts/base.html" if usuario_actual.tipo_estudiante == "administrador" else "layouts/base2.html"

    cursos = [
        "Razonamiento Verbal", "Razonamiento Matemático", "Aritmética", "Álgebra",
        "Geometría", "Trigonometría", "Física", "Química", "Biología", "Lenguaje",
        "Literatura", "Historia", "Geografía", "Filosofía", "Psicología", "Economía"
    ]

    # Cantidad real de preguntas por curso
    preguntas_por_curso = {
        "Razonamiento Verbal": 15,
        "Razonamiento Matemático": 15,
        "Aritmética": 5,
        "Álgebra": 5,
        "Geometría": 4,
        "Trigonometría": 4,
        "Física": 5,
        "Química": 5,
        "Biología": 9,
        "Lenguaje": 6,
        "Literatura": 6,
        "Historia": 3,
        "Geografía": 2,
        "Filosofía": 2,
        "Psicología": 2,
        "Economía": 2,
    }

    ruta_base = r'media'
    graficos = []
    fechas = set()

    for estudiante in Estudiante.objects.all():
        reportes = Reporte.objects.filter(KK_usuario=estudiante).order_by('-fecha_de_examen')

        for reporte in reportes:
            datos = reporte.obtener_datos()
            carpeta_usuario = os.path.join(ruta_base, f"{estudiante.usuario}_2025")
            os.makedirs(carpeta_usuario, exist_ok=True)

            for curso, valores in datos.items():
                correctas, incorrectas = valores
                total_preguntas = preguntas_por_curso.get(curso, 10)  # Si no se encuentra, usa 10 como fallback
                en_blanco = total_preguntas - (correctas + incorrectas)

                fig, ax = plt.subplots(figsize=(8, 4))
                ax.bar(["Correctas", "Incorrectas", "En blanco"],
                       [correctas, incorrectas, en_blanco],
                       color=["green", "red", "gray"])

                ax.set_title(f"{curso} - {reporte.fecha_de_examen}")

                nombre_archivo = f"{curso}_{estudiante.usuario}_{reporte.fecha_de_examen}.png".replace(" ", "_")
                ruta_archivo = os.path.join(carpeta_usuario, nombre_archivo)

                plt.savefig(ruta_archivo, format='png')

                buffer = BytesIO()
                plt.savefig(buffer, format='png')
                buffer.seek(0)
                image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                buffer.close()
                plt.close(fig)

                graficos.append({
                    "fecha": reporte.fecha_de_examen,
                    "curso": curso,
                    "grafico": image_base64
                })

            fechas.add(reporte.fecha_de_examen)

    return render(request, 'ultimo_simulacro.html', {
        'graficos': graficos,
        'cursos': cursos,
        'fechas': sorted(fechas),
        "base_template": base_template
    })


@login_required
@estudiante_tipo_requerido(['administrador'])
def subir_reporte_respaldo(request):
    
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular

    if request.method == 'POST':
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            try:
                # Leer el archivo Excel
                df = pd.read_excel(archivo_excel)

                # Validar que el archivo tenga las columnas correctas
                columnas_esperadas = [
                    'Alumno', 'Fecha Simulacro', 'Puesto', 'Rv_1', 'Rv_2', 'Rm_1', 'Rm_2', 'Ar_1', 'Ar_2', 'Al_1', 'Al_2','Ge_1', 'Ge_2', 'Tr_1', 'Tr_2', 'Fi_1', 'Fi_2', 'Qu_1', 'Qu_2', 'Bi_1', 'Bi_2','Le_1', 'Le_2', 'Lit_1', 'Lit_2', 'Hi_1', 'Hi_2', 'Gf_1', 'Gf_2', 'Fil_1', 'Fil_2','Psi_1', 'Psi_2', 'Ec_1', 'Ec_2', 'Observación'
                ]
                if not all(col in df.columns for col in columnas_esperadas):
                    messages.error(request, "El archivo Excel no tiene las columnas esperadas.")
                    return redirect('subir_reporte')

                # Procesar cada fila del archivo Excel
                for _, fila in df.iterrows():
                    estudiante = Estudiante.objects.get(usuario=fila['Alumno'])
                    Reporte.objects.create(
                        KK_usuario=estudiante,
                        Rv_1=fila['Rv_1'],
                        Rv_2=fila['Rv_2'],
                        Rm_1=fila['Rm_1'],
                        Rm_2=fila['Rm_2'],
                        Ar_1=fila['Ar_1'],
                        Ar_2=fila['Ar_2'],
                        Al_1=fila['Al_1'],
                        Al_2=fila['Al_2'],
                        Ge_1=fila['Ge_1'],
                        Ge_2=fila['Ge_2'],
                        Tr_1=fila['Tr_1'],
                        Tr_2=fila['Tr_2'],
                        Fi_1=fila['Fi_1'],
                        Fi_2=fila['Fi_2'],
                        Qu_1=fila['Qu_1'],
                        Qu_2=fila['Qu_2'],
                        Bi_1=fila['Bi_1'],
                        Bi_2=fila['Bi_2'],
                        Le_1=fila['Le_1'],
                        Le_2=fila['Le_2'],
                        Lit_1=fila['Lit_1'],
                        Lit_2=fila['Lit_2'],
                        Hi_1=fila['Hi_1'],
                        Hi_2=fila['Hi_2'],
                        Gf_1=fila['Gf_1'],
                        Gf_2=fila['Gf_2'],
                        Fil_1=fila['Fil_1'],
                        Fil_2=fila['Fil_2'],
                        Psi_1=fila['Psi_1'],
                        Psi_2=fila['Psi_2'],
                        Ec_1=fila['Ec_1'],
                        Ec_2=fila['Ec_2'],
                        Observacion=fila['Observación'],
                        fecha_de_examen=fila['Fecha Simulacro'],
                        puesto=fila['Puesto']
                    )

                messages.success(request, "Datos cargados exitosamente.")
                return redirect('subir_reporte')

            except Estudiante.DoesNotExist:
                messages.error(request, "El IDEstudiante no existe en la base de datos.")
                return redirect('subir_reporte')
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('subir_reporte')
    else:
        form = CargarExcelForm()
    
    return render(request, "subir_reporte.html", {"base_template": base_template, "form": form})

@login_required
@estudiante_tipo_requerido(['administrador'])
def descargar_reportes_zip_respaldo(request):
    # Ruta a la carpeta de simulacros
    carpeta = os.path.join(settings.MEDIA_ROOT, "reportes", "simulacros")
    
    if not os.path.exists(carpeta):
        raise Http404("La carpeta de reportes no existe")

    # Creamos el archivo ZIP en memoria
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for nombre_archivo in os.listdir(carpeta):
            ruta_archivo = os.path.join(carpeta, nombre_archivo)
            if os.path.isfile(ruta_archivo):
                zip_file.write(ruta_archivo, arcname=nombre_archivo)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="reportes_simulacros.zip"'
    return response

@login_required
@estudiante_tipo_requerido(['administrador'])
def descargar_reportes_asistencia_zip_respaldo(request):
    # Ruta a la carpeta de asistenicias
    carpeta = os.path.join(settings.MEDIA_ROOT, "reportes", "asistencias")
    
    if not os.path.exists(carpeta):
        raise Http404("La carpeta de reportes no existe, comuníquese con soporte")

    # Creamos el archivo ZIP en memoria
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for nombre_archivo in os.listdir(carpeta):
            ruta_archivo = os.path.join(carpeta, nombre_archivo)
            if os.path.isfile(ruta_archivo):
                zip_file.write(ruta_archivo, arcname=nombre_archivo)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="reportes_asistencias.zip"'
    return response


@login_required
@estudiante_tipo_requerido(['administrador'])
def generar_todo_reporte_respaldo_celery(request):
    try:
        estudiantes = Estudiante.objects.filter(reporte_actualizado = True)
        ruta_guardar = 'media/reportes/simulacros'
        plantilla_path = "login/static/img/plantilla-notas.png"

        if not os.path.exists(plantilla_path):
            raise Http404("No se encontró la plantilla del reporte.")

        # Definimos cursos y preguntas por nivel
        
        for estudiante in estudiantes:
            reportes_qs = estudiante.reportes.all().order_by('-fecha_de_examen')

            if not reportes_qs.exists():
                continue

            # Tomar los últimos 8 y ordenarlos cronológicamente
            reportes = list(reportes_qs[:8][::-1])


            ruta_carpeta = f'media/{estudiante.usuario}_2025/imagenes_reporte/'
            os.makedirs(ruta_carpeta, exist_ok=True)

            # 1. Generar gráficos por curso según nivel
            

            # 1. Recolectar datos por curso según nivel
            nivel = reportes[0].nivel # Todos los reportes del estudiante tienen mismo nivel
            if nivel == 30:
                cursos = cursos_pre
                preguntas_por_curso = preguntas_semillero
            else:
                cursos = cursos_pre
                preguntas_por_curso = preguntas_pre

            datos_por_curso = {curso: [] for curso in cursos}

            for reporte in reportes:
                datos_cursos = reporte.obtener_datos()
                fecha = reporte.fecha_de_examen

                for curso in cursos:
                    if curso not in datos_cursos:
                        continue
                    correctas, incorrectas = datos_cursos[curso]
                    total = preguntas_por_curso.get(curso, 10)

                    datos_por_curso[curso].append({
                        "Fecha Simulacro": fecha,
                        "correctas": correctas,
                        "incorrectas": incorrectas,
                        "total": total
                    })

            # 2. Generar gráfico por curso con función externa
            for curso, datos in datos_por_curso.items():
                crear_grafico_estudiante_curso(estudiante, curso, datos)


            # 2. Gráfico de puntajes por fecha
            fechas = [r.fecha_de_examen.strftime('%Y-%m-%d') for r in reportes]
            puntajes = [r.obtener_total_puntaje() for r in reportes]

            def generar_grafico_puntaje_barras(x, y):
                fig, ax = plt.subplots(figsize=(6, 3))
                barras = ax.bar(x, y, color='#DA880B')
                for spine in ax.spines.values():
                    spine.set_visible(False)
                ax.set_title("")
                ax.set_xlabel("")
                ax.set_ylabel("")
                for barra, valor in zip(barras, y):
                    altura = barra.get_height()
                    ax.text(barra.get_x() + barra.get_width() / 2, altura + 1, f'{valor}',
                            ha='center', va='bottom', fontsize=20, fontweight='bold')
                tamaño_letra = 10 if len(x) > 5 else 10
                plt.xticks(rotation=0, fontsize=tamaño_letra)
                ax.tick_params(axis='y', left=False, labelleft=False)
                ax.grid(False)
                ruta_grafico = os.path.join(ruta_carpeta, "zgrafico_puntaje.png")
                fig.tight_layout()
                fig.savefig(ruta_grafico, bbox_inches='tight')
                plt.close(fig)
                return ruta_grafico

            ruta_grafico_puntaje = generar_grafico_puntaje_barras(fechas, puntajes)

            # 3. Crear imagen final con todos los gráficos
            plantilla = Image.open(plantilla_path)
            draw = ImageDraw.Draw(plantilla)
            fuente = ImageFont.truetype("arial.ttf", 40)
            draw.text((720, 124), estudiante.nombre, fill="black", font=fuente)

            a, b, i = 180, 210, 1
            imagenes = os.listdir(ruta_carpeta)
            imagenes = [img for img in imagenes if img.endswith(".png") and img != "zgrafico_puntaje.png"]

            for imagen in sorted(imagenes):
                ruta_imagen = os.path.join(ruta_carpeta, imagen)
                if not os.path.exists(ruta_imagen):
                    continue
                if i % 2 == 0:
                    a = 960
                else:
                    a = 180
                    b += 240
                otra_imagen = Image.open(ruta_imagen)
                imagen_redimensionada = otra_imagen.resize((700, 200))
                plantilla.paste(imagen_redimensionada, (a, b))
                i += 1

            if os.path.exists(ruta_grafico_puntaje):
                grafico_puntaje_img = Image.open(ruta_grafico_puntaje).resize((900, 200))
                plantilla.paste(grafico_puntaje_img, (300, 260))

            os.makedirs(ruta_guardar, exist_ok=True)
            resultado_path = os.path.join(ruta_guardar, f"{estudiante.usuario}_reporte_simulacro.png")
            plantilla.save(resultado_path)

        messages.success(request, "Gráficos y reportes generados exitosamente para todos los estudiantes.")
        return redirect('seleccionar_fecha_generacion')

    except Exception as e:
        print(f"Error al generar todo el reporte: {e}")
        raise Http404("Ocurrió un error al generar los reportes. {e}")

@login_required
@estudiante_tipo_requerido(['administrador'])
def cargar_asistencias_respaldo(request):
    estudiante = request.user
    if estudiante.tipo_estudiante == "administrador":  
        base_template = "layouts/base.html"  # Plantilla para administrador
    else:
        base_template = "layouts/base2.html"  # Plantilla para estudiante regular
    if request.method == 'POST':
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            try:
                df = pd.read_excel(archivo_excel)
                columnas_esperadas = [
                    'Cód. Alumno', 'Fecha', 'Marca 01', 'Marca 02', 'Marca 03', 'Marca 04', 'Marca 05', 'Marca 06', 'Observ'
                ]
                if not all(col in df.columns for col in columnas_esperadas):
                    messages.error(request, "El archivo Excel no tiene las columnas esperadas.")
                    return redirect('subir_asistencia')
                for _, fila in df.iterrows():
                    estudiante = Estudiante.objects.get(usuario=fila['Cód. Alumno'])
                    Asistencia.objects.create(
                        KK_usuario=estudiante,
                        Fecha=fila['Fecha'],
                        Ingreso_mana=fila['Marca 01'],
                        Salida_mana=fila['Marca 02'],
                        Ingreso_tarde=fila['Marca 03'],
                        Salida_tarde=fila['Marca 04'],
                        Ingreso_noche=fila['Marca 05'],
                        Salida_noche=fila['Marca 06'],
                        Observacion=fila['Observ']
                    )
                messages.success(request, "Datos cargados exitosamente.")
                return redirect('subir_asistencia')

            except Estudiante.DoesNotExist:
                messages.error(request, "El usuario no existe en la base de datos.")
                return redirect('subir_asistencia')
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('subir_asistencia')
    else:
        form = CargarExcelForm()
    
    return render(request, "subir_asistencia.html", {"base_template": base_template, "form": form})

@login_required
@estudiante_tipo_requerido(['administrador'])
def cargar_asistencias_respaldo_2(request):
    estudiante = request.user
    base_template = "layouts/base.html" if estudiante.tipo_estudiante == "administrador" else "layouts/base2.html"

    if request.method == 'POST':
        form = CargarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            try:
                df = pd.read_excel(archivo_excel)
                columnas_esperadas = ['usuario', 'marca', 'modalidad']
                if not all(col in df.columns for col in columnas_esperadas):
                    messages.error(request, "El archivo Excel no tiene las columnas esperadas: 'usuario', 'marca' y 'modalidad'")
                    return redirect('subir_asistencia')

                for _, fila in df.iterrows():
                    try:
                        estudiante = Estudiante.objects.get(usuario=fila['usuario'])
                        # Parsear fecha y hora desde la columna 'marca'
                        marca_datetime = pd.to_datetime(fila['marca'])

                        fecha = marca_datetime.date()
                        hora = marca_datetime.strftime("%H:%M")
                        modalidad = fila.get('modalidad')
                        Asistencia.objects.create(
                            KK_usuario=estudiante,
                            Fecha=fecha,
                            Hora=hora,
                            Modalidad = modalidad
                        )
                    except Estudiante.DoesNotExist:
                        messages.warning(request, f"El usuario '{fila['usuario']}' no existe en la base de datos.")
                        continue
                    except Exception as e:
                        messages.warning(request, f"Error procesando fila: {str(e)}")
                        continue

                messages.success(request, "Datos cargados exitosamente.")
                return redirect('subir_asistencia')

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('subir_asistencia')
    else:
        form = CargarExcelForm()

    return render(request, "subir_asistencia.html", {"base_template": base_template, "form": form})

@login_required
def ver_todas_asistencias_respaldo(request):
    fecha_filtrada = request.GET.get('fecha')
    asistencias = Asistencia.objects.all().select_related('KK_usuario').order_by('Fecha', 'Hora')

    if fecha_filtrada:
        try:
            fecha_obj = datetime.strptime(fecha_filtrada, "%Y-%m-%d").date()
            asistencias = asistencias.filter(Fecha=fecha_obj)
        except ValueError:
            pass

    datos = []
    for fecha in asistencias.values_list('Fecha', flat=True).distinct().order_by('Fecha'):
        marcas_fecha = asistencias.filter(Fecha=fecha).order_by('KK_usuario__usuario', 'Hora')
        marcas_por_usuario = defaultdict(list)

        for asistencia in marcas_fecha:
            marcas_por_usuario[asistencia.KK_usuario.usuario].append(asistencia)

        for usuario, marcas in marcas_por_usuario.items():
            for i, asistencia in enumerate(marcas, start=1):
                dia_semana = dias_semana[fecha.weekday()]
                datos.append({
                    "id": asistencia.ID_Reporte,
                    "usuario": asistencia.KK_usuario.usuario,
                    "nombre": asistencia.KK_usuario.nombre,
                    "numero_marca": i,
                    "fecha": fecha.strftime('%d/%m/%Y'),
                    "marca": "ENTRADA" if i in [1, 3] else "SALIDA",
                    "hora": asistencia.Hora,
                    "modalidad": asistencia.Modalidad,
                    "observacion": asistencia.Observacion
                })

    return render(request, "listar_asistencias.html", {
        "asistencias": datos,
        "base_template": "layouts/base.html",  # O ajusta según usuario
        "fecha_filtrada": fecha_filtrada
    })

